import os
import time
from datetime import datetime, timedelta

import pandas as pd

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule


from wms_processor import procesar_archivo
from personal_config import normalizar_nombre
from wms_utils import (
    _ensure_columns,
    _normalize_column_names,
    _parse_fecha_excel,
    _parse_hora_excel,
    _build_fechahora,
    _norm_desc_series,
    _aplicar_estilo_excel_resultados
)


MAX_DIAS_RANGO_PROCESO = 31


SHEETS_CONSOLIDAR = [
    "Trazabilidad",
    "Movimientos Consolidados",
    "Analisis movimientos Bodega",
    "Inactividad Bodega",
    "Analisis cargues Despachos",
    "Inactividad Despacho"
]


COLS_BASE_WMS = [
    "Item Number",
    "Número Lote",
    "Código Transacción",
    "Description",
    "ID Bodega",
    "Desde ID Ubicación",
    "A ID Ubicación",
    "LP",
    "ID Operador",
    "Nombre",
    "Fecha",
    "Tiempo",
    "Cantidad",
    "Tipo"
]


# ============================================================
# FECHAS
# ============================================================

def _parse_fecha_ui(fecha_str):
    fecha = pd.to_datetime(
        fecha_str,
        format="%d/%m/%Y",
        errors="coerce"
    )

    if pd.isna(fecha):
        return None

    return fecha.date()


def generar_fechas_rango(fecha_inicio_str, fecha_fin_str):
    fecha_inicio = _parse_fecha_ui(fecha_inicio_str)
    fecha_fin = _parse_fecha_ui(fecha_fin_str)

    if fecha_inicio is None or fecha_fin is None:
        raise ValueError(
            "Fecha inicio o fecha fin inválida. Usa formato dd/mm/aaaa."
        )

    if fecha_fin < fecha_inicio:
        raise ValueError(
            "La fecha fin no puede ser menor que la fecha inicio."
        )

    dias = (fecha_fin - fecha_inicio).days + 1

    if dias > MAX_DIAS_RANGO_PROCESO:
        raise ValueError(
            f"El rango seleccionado tiene {dias} días. "
            f"El máximo permitido es {MAX_DIAS_RANGO_PROCESO} días."
        )

    return [
        fecha_inicio + timedelta(days=i)
        for i in range(dias)
    ]


# ============================================================
# CARGA ÚNICA DEL EXCEL
# ============================================================

def cargar_excel_origen_preparado(ruta_excel):
    """
    Lee el Excel WMS una sola vez y deja columnas auxiliares listas.

    Este DataFrame será filtrado en memoria por fecha/turno.
    """

    df0 = pd.read_excel(
        ruta_excel,
        engine="openpyxl"
    )

    df0 = _normalize_column_names(df0)

    df0 = _ensure_columns(df0, COLS_BASE_WMS)
    df0 = df0[COLS_BASE_WMS].copy()

    df0["Description_norm"] = _norm_desc_series(df0["Description"])
    df0["Nombre_norm"] = df0["Nombre"].apply(normalizar_nombre)

    df0["Fecha_dt"] = df0["Fecha"].apply(_parse_fecha_excel)
    df0["Hora_dt"] = df0["Tiempo"].apply(_parse_hora_excel)
    df0["FechaHora"] = _build_fechahora(df0)

    df0 = df0[
        df0["Fecha_dt"].notna() &
        df0["Hora_dt"].notna() &
        df0["FechaHora"].notna()
    ].copy()

    df0 = df0.reset_index(drop=True)
    df0["_row_id_wms"] = df0.index

    return df0


def detectar_fechas_df_preparado(df_origen):
    if df_origen is None or df_origen.empty:
        return []

    fechas = (
        df_origen["Fecha_dt"]
        .dropna()
        .drop_duplicates()
        .sort_values()
    )

    return list(fechas)


def filtrar_df_por_turno_preparado(
    df_origen,
    fecha_base,
    hora_inicio,
    hora_fin
):
    """
    Filtra el DataFrame ya leído y preparado por fecha/turno.

    Usa la misma lógica:
    - Turno normal: fecha_base entre inicio y fin.
    - Turno que cruza medianoche: fecha_base representa el cierre del turno.
    """

    if df_origen is None or df_origen.empty:
        return pd.DataFrame(columns=COLS_BASE_WMS)

    try:
        h_ini = datetime.strptime(str(hora_inicio).strip(), "%H:%M").time()
        h_fin = datetime.strptime(str(hora_fin).strip(), "%H:%M").time()
    except Exception:
        return pd.DataFrame(columns=df_origen.columns)

    df = df_origen.copy()

    cruza = h_ini > h_fin

    if not cruza:
        mask = (
            (df["Fecha_dt"] == fecha_base) &
            (df["Hora_dt"] >= h_ini) &
            (df["Hora_dt"] <= h_fin)
        )
    else:
        fecha_turno = fecha_base - timedelta(days=1)

        cond_noche = df["Hora_dt"] >= h_ini
        cond_madrugada = df["Hora_dt"] <= h_fin

        fecha_turno_col = df["Fecha_dt"].copy()

        fecha_turno_col[cond_madrugada] = (
            fecha_turno_col[cond_madrugada] - timedelta(days=1)
        )

        mask = (
            (fecha_turno_col == fecha_turno) &
            (cond_noche | cond_madrugada)
        )

    df_turno = df[mask].copy()
    df_turno = df_turno.reset_index(drop=True)
    df_turno["_row_id_wms"] = df_turno.index

    return df_turno


# ============================================================
# UTILIDADES
# ============================================================

def _insertar_columnas_contexto(df, fecha_str, turno_nombre, archivo_origen):
    if df is None:
        df = pd.DataFrame()

    df = df.copy()

    for col in ["Fecha_Proceso", "Turno_Proceso", "Archivo_Proceso"]:
        if col in df.columns:
            df = df.drop(columns=[col])

    df.insert(0, "Fecha_Proceso", fecha_str)
    df.insert(1, "Turno_Proceso", turno_nombre)
    df.insert(2, "Archivo_Proceso", os.path.basename(str(archivo_origen)))

    return df


def _actualizar_progreso(window, porcentaje, texto):
    try:
        window["-PROG-"].update(int(porcentaje))
        window["-PERC-"].update(f"{int(porcentaje)}%")
        window["-STATUS-"].update(str(texto)[:160])
        window.refresh()
    except Exception:
        pass


# ============================================================
# RESÚMENES MENSUALES
# ============================================================

def _generar_resumen_mensual_desde_movimientos(df_movimientos):
    columnas = [
        "ID Operador",
        "Nombre",
        "Rol",
        "Turnos_Procesados",
        "Movimientos",
        "Cantidad_Total"
    ]

    if df_movimientos is None or df_movimientos.empty:
        return pd.DataFrame(columns=columnas)

    df = df_movimientos.copy()
    
    # Normalizar encabezados por seguridad
    df.columns = [
        str(c).strip()
        for c in df.columns
    ]

    for col in [
        "ID Operador",
        "Nombre",
        "Rol",
        "Cantidad",
        "Movimiento_Global_ID",
        "Turno_Proceso"
    ]:
        if col not in df.columns:
            df[col] = ""

    df["Cantidad"] = pd.to_numeric(
        df["Cantidad"],
        errors="coerce"
    ).fillna(0)

    resumen = (
        df
        .groupby(["ID Operador", "Nombre", "Rol"], dropna=False)
        .agg(
            Turnos_Procesados=("Turno_Proceso", "nunique"),
            Movimientos=("Movimiento_Global_ID", "count"),
            Cantidad_Total=("Cantidad", "sum")
        )
        .reset_index()
        .sort_values(
            ["Cantidad_Total", "Movimientos"],
            ascending=False
        )
    )

    return resumen


def _generar_resumen_por_rol(df_movimientos):
    columnas = [
        "Rol",
        "Operadores",
        "Movimientos",
        "Cantidad_Total"
    ]

    if df_movimientos is None or df_movimientos.empty:
        return pd.DataFrame(columns=columnas)

    df = df_movimientos.copy()

    for col in ["Rol", "Nombre", "Cantidad", "Movimiento_Global_ID"]:
        if col not in df.columns:
            df[col] = ""

    df["Cantidad"] = pd.to_numeric(
        df["Cantidad"],
        errors="coerce"
    ).fillna(0)

    resumen = (
        df
        .groupby(["Rol"], dropna=False)
        .agg(
            Operadores=("Nombre", "nunique"),
            Movimientos=("Movimiento_Global_ID", "count"),
            Cantidad_Total=("Cantidad", "sum")
        )
        .reset_index()
        .sort_values(
            ["Cantidad_Total", "Movimientos"],
            ascending=False
        )
    )

    return resumen


def _generar_resumen_por_turno(df_movimientos):
    columnas = [
        "Fecha_Proceso",
        "Turno_Proceso",
        "Operadores",
        "Movimientos",
        "Cantidad_Total"
    ]

    if df_movimientos is None or df_movimientos.empty:
        return pd.DataFrame(columns=columnas)

    df = df_movimientos.copy()

    for col in [
        "Fecha_Proceso",
        "Turno_Proceso",
        "Nombre",
        "Cantidad",
        "Movimiento_Global_ID"
    ]:
        if col not in df.columns:
            df[col] = ""

    df["Cantidad"] = pd.to_numeric(
        df["Cantidad"],
        errors="coerce"
    ).fillna(0)

    resumen = (
        df
        .groupby(["Fecha_Proceso", "Turno_Proceso"], dropna=False)
        .agg(
            Operadores=("Nombre", "nunique"),
            Movimientos=("Movimiento_Global_ID", "count"),
            Cantidad_Total=("Cantidad", "sum")
        )
        .reset_index()
    )

    return resumen


def _escribir_resumen_estados_visual(writer, df_movimientos):
    """
    Crea la hoja 'Resumen Estados' con formato visual mensual por líder.

    Bloques:
        Lider X
        Nombre | Rol | Total | %

    Agrupación:
        Lider + Rol + Nombre

    Total:
        Número de movimientos productivos del operador.

    %:
        Total operador / máximo Total dentro del mismo Lider + Rol.
    """

    wb = writer.book
    sheet_name = "Resumen Estados"

    if sheet_name in wb.sheetnames:
        ws_old = wb[sheet_name]
        wb.remove(ws_old)

    ws = wb.create_sheet(sheet_name)

    # ============================================================
    # VALIDACIÓN BASE
    # ============================================================

    if df_movimientos is None or df_movimientos.empty:
        ws["A1"] = "No hay movimientos consolidados para resumir."
        ws["A1"].font = Font(bold=True)
        return

    df = df_movimientos.copy()

    for col in [
        "Lider",
        "Nombre",
        "Rol",
        "Movimiento_Global_ID",
        "Tipo_Match"
    ]:
        if col not in df.columns:
            df[col] = ""

    # Evitar contar padres lógicos si todavía aparecen
    df = df[
        df["Tipo_Match"].astype(str).str.upper().str.strip() != "MULTIDESTINO_PADRE"
    ].copy()

    df["Lider"] = (
        df["Lider"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df["Nombre"] = (
        df["Nombre"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df["Rol"] = (
        df["Rol"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df.loc[df["Lider"] == "", "Lider"] = "SIN LIDER"
    df.loc[df["Nombre"] == "", "Nombre"] = "SIN NOMBRE"
    df.loc[df["Rol"] == "", "Rol"] = "SIN ROL"

    # ============================================================
    # TOTAL POR OPERADOR DENTRO DE LIDER + ROL
    # ============================================================

    resumen = (
        df
        .groupby(["Lider", "Rol", "Nombre"], dropna=False)
        .size()
        .reset_index(name="Total")
    )

    if resumen.empty:
        ws["A1"] = "No hay datos para construir el resumen."
        ws["A1"].font = Font(bold=True)
        return

    resumen["Max_Lider_Rol"] = (
        resumen
        .groupby(["Lider", "Rol"])["Total"]
        .transform("max")
    )

    resumen["Porcentaje"] = resumen.apply(
        lambda r: round(r["Total"] / r["Max_Lider_Rol"], 4)
        if r["Max_Lider_Rol"] > 0 else 0,
        axis=1
    )

    resumen = resumen.sort_values(
        ["Lider", "Rol", "Total"],
        ascending=[True, True, False]
    )

    # ============================================================
    # ESTILOS
    # ============================================================

    fill_title = PatternFill(
        start_color="FFFFFF",
        end_color="FFFFFF",
        fill_type="solid"
    )

    fill_header = PatternFill(
        start_color="156082",
        end_color="156082",
        fill_type="solid"
    )

    fill_green = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    fill_red = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    font_title = Font(
        bold=True,
        size=14,
        color="000000"
    )

    font_header = Font(
        bold=True,
        color="FFFFFF",
        size=12
    )

    font_body = Font(
        size=11,
        color="000000"
    )

    border_blue = Border(
        left=Side(style="thin", color="156082"),
        right=Side(style="thin", color="156082"),
        top=Side(style="thin", color="156082"),
        bottom=Side(style="thin", color="156082")
    )

    alignment_center = Alignment(
        horizontal="center",
        vertical="center"
    )

    alignment_left = Alignment(
        horizontal="left",
        vertical="center"
    )

    # ============================================================
    # ESCRIBIR BLOQUES POR LIDER + ROL
    # ============================================================

    fila_actual = 1

    lideres = resumen["Lider"].drop_duplicates().tolist()

    for lider in lideres:
        df_lider = resumen[
            resumen["Lider"] == lider
        ].copy()

        roles = df_lider["Rol"].drop_duplicates().tolist()

        for rol in roles:
            bloque = df_lider[
                df_lider["Rol"] == rol
            ].copy()

            if bloque.empty:
                continue

            # Título del bloque
            ws.merge_cells(
                start_row=fila_actual,
                start_column=1,
                end_row=fila_actual,
                end_column=4
            )

            celda_titulo = ws.cell(
                row=fila_actual,
                column=1,
                value=f"Líder {lider}"
            )

            celda_titulo.font = font_title
            celda_titulo.fill = fill_title
            celda_titulo.alignment = alignment_center

            fila_actual += 1

            # Encabezados
            encabezados = ["Nombre", "Rol", "Total", "%"]

            for col_idx, encabezado in enumerate(encabezados, start=1):
                celda = ws.cell(
                    row=fila_actual,
                    column=col_idx,
                    value=encabezado
                )

                celda.fill = fill_header
                celda.font = font_header
                celda.alignment = alignment_center
                celda.border = border_blue

            fila_header = fila_actual
            fila_actual += 1

            fila_inicio_datos = fila_actual

            for _, row in bloque.iterrows():
                ws.cell(
                    row=fila_actual,
                    column=1,
                    value=row["Nombre"]
                )

                ws.cell(
                    row=fila_actual,
                    column=2,
                    value=row["Rol"]
                )

                ws.cell(
                    row=fila_actual,
                    column=3,
                    value=int(row["Total"])
                )

                ws.cell(
                    row=fila_actual,
                    column=4,
                    value=float(row["Porcentaje"])
                )

                for col_idx in range(1, 5):
                    celda = ws.cell(
                        row=fila_actual,
                        column=col_idx
                    )

                    celda.font = font_body
                    celda.border = border_blue

                    if col_idx == 1:
                        celda.alignment = alignment_left
                    else:
                        celda.alignment = alignment_center

                celda_pct = ws.cell(
                    row=fila_actual,
                    column=4
                )

                celda_pct.number_format = "0%"

                if float(row["Porcentaje"]) >= 1:
                    celda_pct.fill = fill_green
                else:
                    celda_pct.fill = fill_red

                fila_actual += 1

            fila_fin_datos = fila_actual - 1

            # Barra de datos en columna %
            if fila_fin_datos >= fila_inicio_datos:
                rango_pct = f"D{fila_inicio_datos}:D{fila_fin_datos}"

                try:
                    ws.conditional_formatting.add(
                        rango_pct,
                        DataBarRule(
                            start_type="num",
                            start_value=0,
                            end_type="num",
                            end_value=1,
                            color="FF6384",
                            showValue=True
                        )
                    )
                except Exception:
                    pass

            # Espacio entre bloques
            fila_actual += 2

    # ============================================================
    # AJUSTES VISUALES
    # ============================================================

    anchos = {
        "A": 42,
        "B": 26,
        "C": 14,
        "D": 14
    }

    for col, width in anchos.items():
        ws.column_dimensions[col].width = width

    ws.sheet_view.showGridLines = False

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "center",
                vertical="center"
            )

def _generar_resumen_estados_trazabilidad(df_trazabilidad):
    """
    Genera resumen técnico de estados desde la hoja Trazabilidad.

    Ejemplo:
    Estado                         Registros
    MOV_REAL                       1500
    SIN_MATCH                      25
    NO_PRODUCTIVO_RETORNO_ORIGEN   18
    """

    columnas = [
        "Estado",
        "Registros"
    ]

    if df_trazabilidad is None or df_trazabilidad.empty:
        return pd.DataFrame(columns=columnas)

    df = df_trazabilidad.copy()

    if "Estado" not in df.columns:
        return pd.DataFrame(columns=columnas)

    df["Estado"] = (
        df["Estado"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df.loc[df["Estado"] == "", "Estado"] = "SIN_ESTADO"

    resumen = (
        df
        .groupby(["Estado"], dropna=False)
        .size()
        .reset_index(name="Registros")
        .sort_values("Registros", ascending=False)
    )

    return resumen     

def _quitar_filtros_resumen_estados(writer):
    """
    Quita cualquier autofiltro o tabla automática en la hoja Resumen Estados.
    Se ejecuta al final porque _aplicar_estilo_excel_resultados puede volver
    a aplicar filtros.
    """

    try:
        wb = writer.book

        if "Resumen Estados" not in wb.sheetnames:
            return

        ws = wb["Resumen Estados"]

        # Quitar autofiltro
        try:
            ws.auto_filter.ref = None
        except Exception:
            pass

        # Quitar tablas si alguna función de estilo las creó
        try:
            if hasattr(ws, "tables"):
                for table_name in list(ws.tables.keys()):
                    del ws.tables[table_name]
        except Exception:
            pass

    except Exception:
        pass       

            
# ============================================================
# EXPORTACIÓN EXCEL MENSUAL
# ============================================================

def _escribir_excel_mensual(
    ruta_salida_mensual,
    consolidados_por_hoja,
    df_log
):
    df_movimientos = consolidados_por_hoja.get(
        "Movimientos Consolidados",
        pd.DataFrame()
    )

    df_trazabilidad = consolidados_por_hoja.get(
        "Trazabilidad",
        pd.DataFrame()
    )

    resumen_operadores = _generar_resumen_mensual_desde_movimientos(
        df_movimientos
    )

    resumen_roles = _generar_resumen_por_rol(
        df_movimientos
    )

    resumen_turnos = _generar_resumen_por_turno(
        df_movimientos
    )

    resumen_estados_trazabilidad = _generar_resumen_estados_trazabilidad(
        df_trazabilidad
    )

    with pd.ExcelWriter(ruta_salida_mensual, engine="openpyxl") as writer:
        resumen_operadores.to_excel(
            writer,
            sheet_name="Resumen Mensual Operadores",
            index=False
        )

        resumen_roles.to_excel(
            writer,
            sheet_name="Resumen Mensual Roles",
            index=False
        )

        resumen_turnos.to_excel(
            writer,
            sheet_name="Resumen Mensual Turnos",
            index=False
        )

        # Hoja visual como la que solicitaste
        _escribir_resumen_estados_visual(
            writer=writer,
            df_movimientos=df_movimientos
        )

        # Hoja adicional con el resumen técnico de estados WMS
        resumen_estados_trazabilidad.to_excel(
            writer,
            sheet_name="Estados Trazabilidad",
            index=False
        )

        df_log.to_excel(
            writer,
            sheet_name="Log Procesamiento",
            index=False
        )

        for sheet_name, df_sheet in consolidados_por_hoja.items():
            if df_sheet is None:
                df_sheet = pd.DataFrame()

            nombre_hoja = str(sheet_name)[:31]

            df_sheet.to_excel(
                writer,
                sheet_name=nombre_hoja,
                index=False
            )

        try:
            _aplicar_estilo_excel_resultados(writer)
        except Exception:
            pass

        # La hoja Resumen Estados debe quedar sin filtros.
        _quitar_filtros_resumen_estados(writer)

    return ruta_salida_mensual


# ============================================================
# PROCESAMIENTO PRINCIPAL POR RANGO / MES
# ============================================================

def procesar_rango_operacion(
    ruta_excel,
    carpeta_destino,
    fecha_inicio_str,
    fecha_fin_str,
    turnos_activos,
    window,
    conservar_archivos_intermedios=False
):
    """
    Procesa un rango máximo de 31 días y genera UN SOLO Excel mensual.

    Optimizado:
    - Lee el Excel origen una sola vez.
    - Normaliza una sola vez.
    - Filtra cada turno en memoria.
    - No crea carpeta _intermedios_mes.
    - No genera Excel por turno.
    - Consolida todo en un único Excel mensual.
    """

    if not ruta_excel:
        raise ValueError("Selecciona un Excel origen.")

    if not carpeta_destino:
        raise ValueError("Selecciona una carpeta destino.")

    if not os.path.exists(ruta_excel):
        raise ValueError(f"No existe el Excel origen:\n{ruta_excel}")

    if not os.path.exists(carpeta_destino):
        raise ValueError(f"No existe la carpeta destino:\n{carpeta_destino}")

    if not turnos_activos:
        raise ValueError("No hay turnos activos configurados.")

    _actualizar_progreso(
        window,
        1,
        "Leyendo Excel origen una sola vez..."
    )

    df_origen = cargar_excel_origen_preparado(ruta_excel)

    if df_origen.empty:
        raise ValueError(
            "El Excel origen no contiene registros válidos con Fecha, Tiempo y FechaHora."
        )

    fechas_excel = detectar_fechas_df_preparado(df_origen)
    fechas_excel_set = set(fechas_excel)

    fechas_rango = generar_fechas_rango(
        fecha_inicio_str,
        fecha_fin_str
    )

    fechas_validas = [
        f for f in fechas_rango
        if f in fechas_excel_set
    ]

    if not fechas_validas:
        raise ValueError(
            "No se encontraron fechas del rango seleccionado dentro del Excel origen."
        )

    fecha_ini_nombre = fechas_validas[0].strftime("%d-%m-%Y")
    fecha_fin_nombre = fechas_validas[-1].strftime("%d-%m-%Y")

    nombre_salida = (
        f"Productividad_MENSUAL_{fecha_ini_nombre}_a_{fecha_fin_nombre}_RESULTADO.xlsx"
    )

    ruta_salida_mensual = os.path.join(
        carpeta_destino,
        nombre_salida
    )

    consolidados_por_hoja = {
        sheet: []
        for sheet in SHEETS_CONSOLIDAR
    }

    log = []

    total_procesos = len(fechas_validas) * len(turnos_activos)
    proceso_actual = 0

    tiempo_inicio_total = time.perf_counter()

    for fecha in fechas_validas:
        fecha_str = fecha.strftime("%d/%m/%Y")

        for turno in turnos_activos:
            proceso_actual += 1

            nombre_turno = str(turno.get("nombre", "")).strip()
            hora_inicio = turno.get("hora_inicio", "07:00")
            hora_fin = turno.get("hora_fin", "14:00")

            progreso_global = int(
                (proceso_actual / max(total_procesos, 1)) * 100
            )

            df_turno = filtrar_df_por_turno_preparado(
                df_origen=df_origen,
                fecha_base=fecha,
                hora_inicio=hora_inicio,
                hora_fin=hora_fin
            )

            if df_turno.empty:
                log.append({
                    "Fecha": fecha_str,
                    "Turno": nombre_turno,
                    "Estado": "SIN_DATOS_TURNO",
                    "Segundos": 0,
                    "Filas_Turno": 0,
                    "Filas_Trazabilidad": 0,
                    "Filas_Movimientos": 0,
                    "Filas_Analisis_Bodega": 0,
                    "Archivo_Intermedio": "",
                    "Error": "No hay registros en el Excel para esta fecha y turno."
                })

                _actualizar_progreso(
                    window,
                    progreso_global,
                    f"Sin datos: {fecha_str} - {nombre_turno} "
                    f"({proceso_actual}/{total_procesos})"
                )

                continue

            _actualizar_progreso(
                window,
                progreso_global,
                f"Procesando: {fecha_str} - {nombre_turno} "
                f"({proceso_actual}/{total_procesos}) | Filas: {len(df_turno)}"
            )

            tiempo_inicio_turno = time.perf_counter()

            try:
                resultado_turno = procesar_archivo(
                    ruta=ruta_excel,
                    carpeta=carpeta_destino,
                    ini=hora_inicio,
                    fin=hora_fin,
                    fecha_str=fecha_str,
                    window=window,
                    turno_config=turno,
                    exportar_excel=False,
                    retornar_dataframes=True,
                    df_pre_filtrado=df_turno,
                    validar_fecha_excel=False,
                    ignorar_faltantes_maestro=True
                )

                if not isinstance(resultado_turno, dict):
                    raise ValueError(
                        "procesar_archivo no devolvió un diccionario. "
                        "Verifica que retornar_dataframes=True esté implementado "
                        "en wms_processor.py."
                    )

                hojas_turno = {
                    "Trazabilidad": resultado_turno.get(
                        "trazabilidad",
                        pd.DataFrame()
                    ),
                    "Movimientos Consolidados": resultado_turno.get(
                        "movimientos",
                        pd.DataFrame()
                    ),
                    "Analisis movimientos Bodega": resultado_turno.get(
                        "analisis_horas_bodega",
                        pd.DataFrame()
                    ),
                    "Inactividad Bodega": resultado_turno.get(
                        "inactividad_bodega",
                        pd.DataFrame()
                    ),
                    "Analisis cargues Despachos": resultado_turno.get(
                        "analisis_horas_despachos",
                        pd.DataFrame()
                    ),
                    "Inactividad Despacho": resultado_turno.get(
                        "inactividad_despacho",
                        pd.DataFrame()
                    )
                }

                df_trazabilidad_tmp = hojas_turno.get(
                    "Trazabilidad",
                    pd.DataFrame()
                )

                df_movimientos_tmp = hojas_turno.get(
                    "Movimientos Consolidados",
                    pd.DataFrame()
                )

                df_analisis_bodega_tmp = hojas_turno.get(
                    "Analisis movimientos Bodega",
                    pd.DataFrame()
                )

                filas_trazabilidad = (
                    len(df_trazabilidad_tmp)
                    if df_trazabilidad_tmp is not None
                    else 0
                )

                filas_movimientos = (
                    len(df_movimientos_tmp)
                    if df_movimientos_tmp is not None
                    else 0
                )

                filas_analisis_bodega = (
                    len(df_analisis_bodega_tmp)
                    if df_analisis_bodega_tmp is not None
                    else 0
                )

                for sheet, df_sheet in hojas_turno.items():
                    if sheet not in consolidados_por_hoja:
                        continue

                    df_sheet = _insertar_columnas_contexto(
                        df_sheet,
                        fecha_str=fecha_str,
                        turno_nombre=nombre_turno,
                        archivo_origen=ruta_excel
                    )

                    consolidados_por_hoja[sheet].append(df_sheet)

                segundos_turno = round(
                    time.perf_counter() - tiempo_inicio_turno,
                    2
                )

                log.append({
                    "Fecha": fecha_str,
                    "Turno": nombre_turno,
                    "Estado": "OK",
                    "Segundos": segundos_turno,
                    "Filas_Turno": len(df_turno),
                    "Filas_Trazabilidad": filas_trazabilidad,
                    "Filas_Movimientos": filas_movimientos,
                    "Filas_Analisis_Bodega": filas_analisis_bodega,
                    "Archivo_Intermedio": "",
                    "Error": ""
                })

            except Exception as e:
                segundos_turno = round(
                    time.perf_counter() - tiempo_inicio_turno,
                    2
                )

                log.append({
                    "Fecha": fecha_str,
                    "Turno": nombre_turno,
                    "Estado": "ERROR",
                    "Segundos": segundos_turno,
                    "Filas_Turno": len(df_turno),
                    "Filas_Trazabilidad": 0,
                    "Filas_Movimientos": 0,
                    "Filas_Analisis_Bodega": 0,
                    "Archivo_Intermedio": "",
                    "Error": str(e)
                })

                continue

    _actualizar_progreso(
        window,
        98,
        "Consolidando hojas mensuales..."
    )

    consolidados_finales = {}

    for sheet, lista_df in consolidados_por_hoja.items():
        if lista_df:
            consolidados_finales[sheet] = pd.concat(
                lista_df,
                ignore_index=True
            )
        else:
            consolidados_finales[sheet] = pd.DataFrame()

    df_log = pd.DataFrame(log)

    ruta_final = _escribir_excel_mensual(
        ruta_salida_mensual=ruta_salida_mensual,
        consolidados_por_hoja=consolidados_finales,
        df_log=df_log
    )

    tiempo_total = round(
        time.perf_counter() - tiempo_inicio_total,
        2
    )

    errores = [
        x for x in log
        if x.get("Estado") == "ERROR"
    ]

    _actualizar_progreso(
        window,
        100,
        f"Procesamiento mensual finalizado en {tiempo_total} segundos."
    )

    return {
        "archivo_mensual": ruta_final,
        "total_procesos": total_procesos,
        "procesados_ok": len([x for x in log if x.get("Estado") == "OK"]),
        "sin_datos": len([x for x in log if x.get("Estado") == "SIN_DATOS_TURNO"]),
        "errores": errores,
        "segundos_total": tiempo_total,
        "log": df_log,
        "carpeta_intermedia": "",
        "reportes_mensuales": [ruta_final]
    }
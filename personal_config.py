import os
import sys
import pandas as pd
import unicodedata
import re
from difflib import SequenceMatcher

# ============================================================
# RUTA DEL MAESTRO DE PERSONAL
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

RUTA_MAESTRO_PERSONAL = os.path.join(
    BASE_DIR,
    "maestro_personal.xlsx"
)

COLUMNAS_MAESTRO = [
    "ID Operador",
    "Nombre",
    "Area",
    "Rol",
    "Lider",
    "Activo",
    "Observacion"
]

AREAS_VALIDAS = {"Bodega", "Despachos", "Mixto", "Excluir"}


def obtener_directorio_app():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)

    return os.path.dirname(os.path.abspath(__file__))


def obtener_ruta_maestro_personal():
    return os.path.join(obtener_directorio_app(), "maestro_personal.xlsx")


def _quitar_tildes(texto):
    texto = str(texto)

    return "".join(
        c for c in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(c)
    )


def normalizar_nombre(valor):
    texto = str(valor).strip()

    if not texto or texto.lower() == "nan":
        return ""

    texto = _quitar_tildes(texto)
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9ñ\s]", " ", texto)
    texto = " ".join(texto.split())

    return texto

def normalizar_id_operador(valor):
    """
    Normaliza el ID Operador para comparación:
    - texto limpio
    - mayúsculas
    - sin espacios dobles
    """
    texto = str(valor).strip()

    if not texto or texto.lower() == "nan" or texto == "<NA>":
        return ""

    texto = " ".join(texto.split())
    texto = texto.upper()

    return texto

def tokens_nombre(valor):
    nombre_norm = normalizar_nombre(valor)

    if not nombre_norm:
        return set()

    return set(nombre_norm.split())


def clave_tokens_ordenados(valor):
    toks = sorted(tokens_nombre(valor))
    return " ".join(toks)


def _similitud(a, b):
    a = str(a).strip()
    b = str(b).strip()

    if not a or not b:
        return 0.0

    return SequenceMatcher(None, a, b).ratio()


def _normalizar_activo(valor):
    texto = str(valor).strip().lower()

    if texto in ("true", "1", "si", "sí", "s", "activo", "yes", "y"):
        return True

    if texto in ("false", "0", "no", "n", "inactivo"):
        return False

    return True


def _normalizar_area(valor):
    texto = str(valor).strip().lower()

    if texto == "bodega":
        return "Bodega"

    if texto == "despachos":
        return "Despachos"

    if texto == "mixto":
        return "Mixto"

    if texto == "excluir":
        return "Excluir"

    return str(valor).strip()

def normalizar_id_operador(valor):
    """
    Normaliza el ID Operador para comparación y visualización.
    """

    if valor is None:
        return ""

    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass

    texto = str(valor).strip()

    if not texto or texto.lower() in ["nan", "none", "<na>"]:
        return ""

    texto = " ".join(texto.split())
    texto = texto.upper()

    return texto

def normalizar_columnas_maestro_personal(df):
    """
    Normaliza nombres de columnas del maestro de personal.

    Acepta variantes como:
    - ID_Operador
    - ID operador
    - Id Operador
    - id_operador

    Y las convierte a:
    - ID Operador
    """

    if df is None:
        return df

    df = df.copy()

    ren = {}

    for col in df.columns:
        col_original = col
        col_norm = str(col).strip().lower()

        col_norm_simple = (
            col_norm
            .replace("_", " ")
            .replace("-", " ")
            .replace(".", " ")
        )

        col_norm_simple = " ".join(col_norm_simple.split())

        if col_norm_simple in [
            "id operador",
            "id operario",
            "id operator",
            "operador id"
        ]:
            ren[col_original] = "ID Operador"

        elif col_norm_simple in ["nombre", "nombres"]:
            ren[col_original] = "Nombre"

        elif col_norm_simple in ["area", "área"]:
            ren[col_original] = "Area"

        elif col_norm_simple in ["activo", "activa"]:
            ren[col_original] = "Activo"
            
        elif col_norm_simple in ["rol", "cargo", "funcion", "función"]:
            ren[col_original] = "Rol"
            
        elif col_norm_simple in ["lider", "líder", "jefe", "supervisor", "leader"]:
            ren[col_original] = "Lider"

        elif col_norm_simple in [
            "observacion",
            "observación",
            "obs",
            "comentario",
            "comentarios"
        ]:
            ren[col_original] = "Observacion"

    if ren:
        df = df.rename(columns=ren)

    return df


def guardar_maestro_personal(df_maestro):
    """
    Guarda maestro_personal.xlsx en formato nuevo:

    Hoja Bodega
    Hoja Despachos

    Conserva:
    ID Operador, Nombre, Rol, Activo, Observacion

    La columna Area se usa internamente para decidir en qué hoja guardar.
    """

    ruta = obtener_ruta_maestro_personal()

    df_guardar = df_maestro.copy()
    df_guardar = normalizar_columnas_maestro_personal(df_guardar)

    for col in COLUMNAS_MAESTRO:
        if col not in df_guardar.columns:
            df_guardar[col] = ""

    df_guardar["Area"] = df_guardar["Area"].apply(_normalizar_area)

    columnas_hoja = [
        "ID Operador",
        "Nombre",
        "Rol",
        "Lider",
        "Activo",
        "Observacion"
    ]

    df_bodega = df_guardar[
        df_guardar["Area"].astype(str).str.strip().eq("Bodega")
    ][columnas_hoja].copy()

    df_despachos = df_guardar[
        df_guardar["Area"].astype(str).str.strip().eq("Despachos")
    ][columnas_hoja].copy()

    # Los mixtos pueden aparecer en ambas hojas si así lo deseas
    df_mixto = df_guardar[
        df_guardar["Area"].astype(str).str.strip().eq("Mixto")
    ][columnas_hoja].copy()

    if not df_mixto.empty:
        df_bodega = pd.concat([df_bodega, df_mixto], ignore_index=True)
        df_despachos = pd.concat([df_despachos, df_mixto], ignore_index=True)

    with pd.ExcelWriter(ruta, engine="openpyxl", mode="w") as writer:
        df_bodega.to_excel(writer, sheet_name="Bodega", index=False)
        df_despachos.to_excel(writer, sheet_name="Despachos", index=False)

    return ruta



def crear_maestro_personal_si_no_existe():
    ruta = obtener_ruta_maestro_personal()

    if os.path.exists(ruta):
        return ruta

    df = pd.DataFrame(columns=COLUMNAS_MAESTRO)

    with pd.ExcelWriter(ruta, engine="openpyxl", mode="w") as writer:
        pd.DataFrame(columns=[
            "ID Operador",
            "Nombre",
            "Rol",
            "Lider",
            "Activo",
            "Observacion"
        ]).to_excel(writer, sheet_name="Bodega", index=False)

        pd.DataFrame(columns=[
            "ID Operador",
            "Nombre",
            "Rol",
            "Lider",
            "Activo",
            "Observacion"
        ]).to_excel(writer, sheet_name="Despachos", index=False)

    return ruta


def migrar_maestro_personal_si_es_necesario():
    ruta = obtener_ruta_maestro_personal()

    if not os.path.exists(ruta):
        crear_maestro_personal_si_no_existe()
        return ruta

    try:
        df_original = pd.read_excel(ruta, sheet_name="Personal", engine="openpyxl")
    except Exception as e:
        raise ValueError(
            "No se pudo leer el archivo maestro_personal.xlsx para migrarlo.\n\n"
            f"Detalle: {str(e)}"
        )

    df_nuevo = pd.DataFrame()

    df_nuevo["Nombre"] = df_original["Nombre"] if "Nombre" in df_original.columns else pd.NA
    df_nuevo["Area"] = df_original["Area"] if "Area" in df_original.columns else pd.NA
    df_nuevo["Activo"] = df_original["Activo"] if "Activo" in df_original.columns else True
    df_nuevo["Observacion"] = df_original["Observacion"] if "Observacion" in df_original.columns else ""

    df_nuevo["Nombre"] = df_nuevo["Nombre"].astype(str).str.strip()
    df_nuevo["Area"] = df_nuevo["Area"].apply(_normalizar_area)
    df_nuevo["Activo"] = df_nuevo["Activo"].apply(_normalizar_activo)
    df_nuevo["Observacion"] = df_nuevo["Observacion"].astype(str).str.strip()

    df_nuevo = df_nuevo[
        (df_nuevo["Nombre"].astype(str).str.strip() != "") &
        (df_nuevo["Nombre"].astype(str).str.strip().str.lower() != "nan")
    ].copy()

    guardar_maestro_personal(df_nuevo)

    return ruta



def cargar_maestro_personal():
    """
    Carga maestro_personal.xlsx.

    Formato nuevo:
        Hoja Bodega
        Hoja Despachos

    Formato anterior compatible:
        Hoja Personal con columna Area
    """

    import os
    import pandas as pd

    ruta = obtener_ruta_maestro_personal()

    if not os.path.exists(ruta):
        return pd.DataFrame(columns=[
            "ID Operador",
            "Nombre",
            "Area",
            "Rol",
            "Lider",
            "Activo",
            "Observacion",
            "Nombre_norm"
        ])

    hojas_objetivo = {
        "Bodega": "Bodega",
        "Despachos": "Despachos"
    }

    try:
        with pd.ExcelFile(ruta, engine="openpyxl") as xls:
            hojas = xls.sheet_names

            dataframes = []

            hojas_detectadas = [
                hoja for hoja in hojas
                if str(hoja).strip() in hojas_objetivo
            ]

            if hojas_detectadas:
                for hoja in hojas_detectadas:
                    try:
                        df_hoja = pd.read_excel(
                            xls,
                            sheet_name=hoja
                        )
                    except Exception:
                        continue

                    if df_hoja is None or df_hoja.empty:
                        continue

                    df_hoja = df_hoja.copy()
                    df_hoja.columns = [
                        str(c).strip()
                        for c in df_hoja.columns
                    ]

                    df_hoja["Area"] = hojas_objetivo[str(hoja).strip()]
                    dataframes.append(df_hoja)

                if dataframes:
                    df = pd.concat(dataframes, ignore_index=True)
                else:
                    df = pd.DataFrame()

            else:
                try:
                    df = pd.read_excel(
                        xls,
                        sheet_name=0
                    )
                except Exception:
                    df = pd.DataFrame()

    except Exception:
        return pd.DataFrame()
    

    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "ID Operador",
            "Nombre",
            "Area",
            "Rol",
            "Lider",
            "Activo",
            "Observacion",
            "Nombre_norm"
        ])

    df = df.copy()
    df.columns = [
        str(c).strip()
        for c in df.columns
    ]

    if "Área" in df.columns and "Area" not in df.columns:
        df["Area"] = df["Área"]

    columnas_base = [
        "ID Operador",
        "Nombre",
        "Area",
        "Rol",
        "Lider",
        "Activo",
        "Observacion"
    ]

    for col in columnas_base:
        if col not in df.columns:
            df[col] = ""

    df["ID Operador"] = df["ID Operador"].astype(str).str.strip()
    df["Nombre"] = df["Nombre"].astype(str).str.strip()
    df["Area"] = df["Area"].astype(str).str.strip()
    df["Rol"] = df["Rol"].astype(str).str.strip()
    df["Lider"] = df["Lider"].astype(str).str.strip()
    df["Activo"] = df["Activo"].astype(str).str.strip()
    df["Observacion"] = df["Observacion"].astype(str).str.strip()

    df["Nombre_norm"] = df["Nombre"].apply(normalizar_nombre)

    return df

def convertir_maestro_a_dos_hojas():
    """
    Convierte maestro_personal.xlsx actual a formato nuevo:

    - Hoja Bodega
    - Hoja Despachos

    Conserva:
    ID Operador, Nombre, Rol, Activo, Observacion.
    """

    import os
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ruta = obtener_ruta_maestro_personal()

    if not os.path.exists(ruta):
        raise FileNotFoundError("No existe maestro_personal.xlsx")

    try:
        df = pd.read_excel(
            ruta,
            sheet_name=0,
            engine="openpyxl"
        )
    except Exception as e:
        raise ValueError(f"No se pudo leer maestro_personal.xlsx: {e}")

    df = df.copy()
    df.columns = [
        str(c).strip()
        for c in df.columns
    ]

    if "Área" in df.columns and "Area" not in df.columns:
        df["Area"] = df["Área"]

    columnas_necesarias = [
        "ID Operador",
        "Nombre",
        "Area",
        "Lider",
        "Rol",
        "Activo",
        "Observacion"
    ]

    for col in columnas_necesarias:
        if col not in df.columns:
            df[col] = ""

    df["Area"] = df["Area"].astype(str).str.strip()

    columnas_salida = [
        "ID Operador",
        "Nombre",
        "Rol",
        "Lider",
        "Activo",
        "Observacion"
    ]

    df_bodega = df[
        df["Area"].astype(str).str.strip().str.lower() == "bodega"
    ][columnas_salida].copy()

    df_despachos = df[
        df["Area"].astype(str).str.strip().str.lower().isin([
            "despachos",
            "despacho"
        ])
    ][columnas_salida].copy()

    # Si no hay despachos todavía, crear hoja vacía con encabezados
    if df_despachos.empty:
        df_despachos = pd.DataFrame(columns=columnas_salida)

    with pd.ExcelWriter(
        ruta,
        engine="openpyxl",
        mode="w"
    ) as writer:
        df_bodega.to_excel(
            writer,
            sheet_name="Bodega",
            index=False
        )

        df_despachos.to_excel(
            writer,
            sheet_name="Despachos",
            index=False
        )

    # ============================================================
    # Estilo básico
    # ============================================================

    wb = load_workbook(ruta)

    fill_header = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    font_header = Font(
        color="FFFFFF",
        bold=True
    )

    border = Border(
        left=Side(style="thin", color="D9EAD3"),
        right=Side(style="thin", color="D9EAD3"),
        top=Side(style="thin", color="D9EAD3"),
        bottom=Side(style="thin", color="D9EAD3")
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        for col_idx, col in enumerate(ws.columns, start=1):
            max_length = 0

            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(max_length + 2, 12),
                45
            )

    wb.save(ruta)

    return True




def validar_maestro_personal(df_maestro):
    faltantes = [c for c in COLUMNAS_MAESTRO if c not in df_maestro.columns]

    if faltantes:
        return False, (
            "El maestro de personal no tiene todas las columnas requeridas.\n\n"
            f"Faltan: {', '.join(faltantes)}"
        )

    if df_maestro.empty:
        return False, (
            "El maestro de personal está vacío.\n\n"
            "Debes registrar al menos una persona antes de procesar."
        )

    df_validar = df_maestro.copy()

    if "Nombre_norm" not in df_validar.columns:
        df_validar["Nombre_norm"] = df_validar["Nombre"].apply(normalizar_nombre)

    duplicados_norm = (
        df_validar["Nombre_norm"][df_validar["Nombre_norm"].duplicated()]
        .dropna()
        .tolist()
    )

    duplicados_norm = sorted(set([x for x in duplicados_norm if x]))

    if duplicados_norm:
        df_dup = df_validar[df_validar["Nombre_norm"].isin(duplicados_norm)].copy()
        nombres_dup = df_dup["Nombre"].astype(str).tolist()

        return False, (
            "El maestro de personal tiene nombres duplicados "
            "(después de normalizar mayúsculas, tildes y espacios).\n\n"
            "Nombres detectados:\n- " + "\n- ".join(nombres_dup)
        )

    areas_invalidas = (
        df_validar.loc[~df_validar["Area"].isin(AREAS_VALIDAS), "Area"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    if areas_invalidas:
        return False, (
            "El maestro de personal tiene áreas inválidas.\n\n"
            "Áreas permitidas: Bodega, Despachos, Mixto, Excluir\n\n"
            "Áreas inválidas detectadas:\n- " + "\n- ".join(areas_invalidas)
        )

    return True, "Maestro de personal válido."


def _preparar_maestro_para_matching(df_maestro):
    df = df_maestro.copy()

    if "ID Operador" not in df.columns:
        df["ID Operador"] = ""

    if "ID_Operador_norm" not in df.columns:
        df["ID_Operador_norm"] = df["ID Operador"].apply(normalizar_id_operador)

    if "Nombre_norm" not in df.columns:
        df["Nombre_norm"] = df["Nombre"].apply(normalizar_nombre)

    if "Nombre_tokens_key" not in df.columns:
        df["Nombre_tokens_key"] = df["Nombre"].apply(clave_tokens_ordenados)

    df["Tokens"] = df["Nombre"].apply(tokens_nombre)

    return df


def resolver_nombre_en_maestro(nombre_turno, df_maestro):
    df = _preparar_maestro_para_matching(df_maestro)

    nombre_norm_turno = normalizar_nombre(nombre_turno)
    tokens_turno = tokens_nombre(nombre_turno)
    tokens_key_turno = clave_tokens_ordenados(nombre_turno)

    resultado_base = {
        "encontrado": False,
        "nombre_norm_turno": nombre_norm_turno,
        "nombre_maestro": "",
        "nombre_norm_maestro": "",
        "area": "",
        "activo": True,
        "metodo": "NO_MATCH",
        "score": 0.0
    }

    if not nombre_norm_turno:
        return resultado_base

    exactos = df[df["Nombre_norm"] == nombre_norm_turno].copy()

    if len(exactos) == 1:
        row = exactos.iloc[0]
        return {
            "encontrado": True,
            "nombre_norm_turno": nombre_norm_turno,
            "nombre_maestro": str(row["Nombre"]),
            "nombre_norm_maestro": str(row["Nombre_norm"]),
            "area": str(row["Area"]),
            "activo": bool(row["Activo"]),
            "metodo": "EXACTO",
            "score": 1.0
        }

    tokens_exactos = df[df["Nombre_tokens_key"] == tokens_key_turno].copy()

    if len(tokens_exactos) == 1:
        row = tokens_exactos.iloc[0]
        return {
            "encontrado": True,
            "nombre_norm_turno": nombre_norm_turno,
            "nombre_maestro": str(row["Nombre"]),
            "nombre_norm_maestro": str(row["Nombre_norm"]),
            "area": str(row["Area"]),
            "activo": bool(row["Activo"]),
            "metodo": "TOKENS_EXACTOS",
            "score": 1.0
        }

    candidatos_subset = []

    if len(tokens_turno) >= 2:
        for _, row in df.iterrows():
            tokens_maestro = row["Tokens"]

            if not tokens_maestro:
                continue

            interseccion = tokens_turno.intersection(tokens_maestro)

            if tokens_turno.issubset(tokens_maestro):
                candidatos_subset.append((row, 1.0))
                continue

            ratio_tokens = len(interseccion) / max(len(tokens_turno), 1)

            if ratio_tokens >= 0.80 and len(interseccion) >= 2:
                candidatos_subset.append((row, ratio_tokens))

    if len(candidatos_subset) == 1:
        row, score = candidatos_subset[0]
        return {
            "encontrado": True,
            "nombre_norm_turno": nombre_norm_turno,
            "nombre_maestro": str(row["Nombre"]),
            "nombre_norm_maestro": str(row["Nombre_norm"]),
            "area": str(row["Area"]),
            "activo": bool(row["Activo"]),
            "metodo": "TOKENS_SUBSET",
            "score": float(score)
        }

    if len(candidatos_subset) > 1:
        candidatos_subset = sorted(candidatos_subset, key=lambda x: x[1], reverse=True)

        mejor_row, mejor_score = candidatos_subset[0]
        segundo_score = candidatos_subset[1][1]

        if mejor_score >= 0.95 and (mejor_score - segundo_score) >= 0.15:
            return {
                "encontrado": True,
                "nombre_norm_turno": nombre_norm_turno,
                "nombre_maestro": str(mejor_row["Nombre"]),
                "nombre_norm_maestro": str(mejor_row["Nombre_norm"]),
                "area": str(mejor_row["Area"]),
                "activo": bool(mejor_row["Activo"]),
                "metodo": "TOKENS_SUBSET",
                "score": float(mejor_score)
            }

    candidatos_fuzzy = []

    for _, row in df.iterrows():
        score = _similitud(tokens_key_turno, str(row["Nombre_tokens_key"]))

        if score >= 0.86:
            candidatos_fuzzy.append((row, score))

    if len(candidatos_fuzzy) == 1:
        row, score = candidatos_fuzzy[0]
        return {
            "encontrado": True,
            "nombre_norm_turno": nombre_norm_turno,
            "nombre_maestro": str(row["Nombre"]),
            "nombre_norm_maestro": str(row["Nombre_norm"]),
            "area": str(row["Area"]),
            "activo": bool(row["Activo"]),
            "metodo": "FUZZY",
            "score": float(score)
        }

    if len(candidatos_fuzzy) > 1:
        candidatos_fuzzy = sorted(candidatos_fuzzy, key=lambda x: x[1], reverse=True)

        mejor_row, mejor_score = candidatos_fuzzy[0]
        segundo_score = candidatos_fuzzy[1][1]

        if mejor_score >= 0.90 and (mejor_score - segundo_score) >= 0.05:
            return {
                "encontrado": True,
                "nombre_norm_turno": nombre_norm_turno,
                "nombre_maestro": str(mejor_row["Nombre"]),
                "nombre_norm_maestro": str(mejor_row["Nombre_norm"]),
                "area": str(mejor_row["Area"]),
                "activo": bool(mejor_row["Activo"]),
                "metodo": "FUZZY",
                "score": float(mejor_score)
            }

    return resultado_base

def resolver_persona_en_maestro(id_operador_turno, nombre_turno, df_maestro):
    """
    Resuelve una persona del Excel origen contra el maestro.

    Prioridad:
    1. ID Operador exacto.
    2. Nombre inteligente.
    """

    df = _preparar_maestro_para_matching(df_maestro)

    id_norm_turno = normalizar_id_operador(id_operador_turno)

    if id_norm_turno:
        coincidencias_id = df[df["ID_Operador_norm"] == id_norm_turno].copy()

        if len(coincidencias_id) == 1:
            row = coincidencias_id.iloc[0]

            return {
                "encontrado": True,
                "id_operador_turno": id_norm_turno,
                "nombre_norm_turno": normalizar_nombre(nombre_turno),
                "id_operador_maestro": str(row.get("ID Operador", "")),
                "nombre_maestro": str(row["Nombre"]),
                "nombre_norm_maestro": str(row["Nombre_norm"]),
                "area": str(row["Area"]),
                "activo": bool(row["Activo"]),
                "metodo": "ID_OPERADOR",
                "score": 1.0
            }

    # Si no encuentra por ID, intenta por nombre inteligente.
    res_nombre = resolver_nombre_en_maestro(nombre_turno, df_maestro)

    res_nombre["id_operador_turno"] = id_norm_turno
    res_nombre["id_operador_maestro"] = ""

    return res_nombre



def construir_asignacion_operadores_para_turno_df(df_turno, df_maestro):
    """
    Construye asignación de área para el personal detectado en el Excel origen.

    Usa:
    1. ID Operador si existe.
    2. Nombre inteligente si no encuentra por ID.

    Retorna:
    {
        "asignacion": {
            nombre_norm_turno: area
        },
        "faltantes": [...],
        "faltantes_raw": [...],
        "matches": [...]
    }
    """

    asignacion = {}
    faltantes = []
    faltantes_raw = []
    matches = []

    if df_turno is None or df_turno.empty:
        return {
            "asignacion": asignacion,
            "faltantes": faltantes,
            "faltantes_raw": faltantes_raw,
            "matches": matches
        }

    df = df_turno.copy()

    if "ID Operador" not in df.columns:
        df["ID Operador"] = ""

    if "Nombre" not in df.columns:
        df["Nombre"] = ""

    df["ID_Operador_norm"] = df["ID Operador"].apply(normalizar_id_operador)
    df["Nombre_norm"] = df["Nombre"].apply(normalizar_nombre)

    df = df[
        (df["Nombre_norm"].astype(str).str.strip() != "") |
        (df["ID_Operador_norm"].astype(str).str.strip() != "")
    ].copy()

    df = df.drop_duplicates(subset=["ID_Operador_norm", "Nombre_norm"]).copy()

    vistos = set()

    for _, row_turno in df.iterrows():
        id_raw = str(row_turno.get("ID Operador", "")).strip()
        nombre_raw = str(row_turno.get("Nombre", "")).strip()

        id_norm = normalizar_id_operador(id_raw)
        nombre_norm = normalizar_nombre(nombre_raw)

        clave_visto = f"{id_norm}|{nombre_norm}"

        if clave_visto in vistos:
            continue

        vistos.add(clave_visto)

        res = resolver_persona_en_maestro(
            id_operador_turno=id_raw,
            nombre_turno=nombre_raw,
            df_maestro=df_maestro
        )

        nombre_norm_turno = res.get("nombre_norm_turno", nombre_norm)

        if res["encontrado"]:
            area = res["area"]

            if not bool(res["activo"]):
                area = "Excluir"

            if nombre_norm_turno:
                asignacion[nombre_norm_turno] = area

            matches.append(res)
        else:
            if nombre_norm_turno:
                faltantes.append(nombre_norm_turno)

            if id_raw:
                faltantes_raw.append(f"{id_raw} - {nombre_raw}")
            else:
                faltantes_raw.append(nombre_raw)

    return {
        "asignacion": asignacion,
        "faltantes": faltantes,
        "faltantes_raw": faltantes_raw,
        "matches": matches
    }

def actualizar_ids_operador_maestro_desde_matches(df_maestro, matches):
    """
    Actualización automática de IDs desactivada.

    Motivo:
    El maestro_personal.xlsx se edita manualmente desde Excel.
    Para evitar bloqueos e infracciones de uso compartido, el software
    no debe guardar ni modificar automáticamente el maestro mientras está abierto.
    """

    return {
        "actualizados": 0,
        "conflictos": [],
        "mensaje": "Actualización automática de IDs desactivada. El maestro se edita manualmente en Excel."
    }
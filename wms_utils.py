import pandas as pd
from datetime import datetime, timedelta, date, time


# ============================================================
# UTILIDADES GENERALES
# ============================================================

def _ensure_columns(df, cols):
    """
    Asegura que el DataFrame tenga todas las columnas indicadas.
    Si falta alguna columna, la crea con pd.NA.
    """
    if df is None:
        raise ValueError("Error interno: _ensure_columns recibió df=None.")

    for c in cols:
        if c not in df.columns:
            df[c] = pd.NA

    return df


def _normalize_column_names(df):
    """
    Normaliza nombres de columnas del Excel origen para evitar errores
    por diferencias de tildes, mayúsculas o nombres alternativos.
    Siempre retorna un DataFrame.
    """
    if df is None:
        raise ValueError("Error interno: _normalize_column_names recibió df=None.")

    ren = {}

    # Código Transacción
    if "Codigo Transacción" in df.columns and "Código Transacción" not in df.columns:
        ren["Codigo Transacción"] = "Código Transacción"

    if "Codigo Transaccion" in df.columns and "Código Transacción" not in df.columns:
        ren["Codigo Transaccion"] = "Código Transacción"

    if "Código Transaccion" in df.columns and "Código Transacción" not in df.columns:
        ren["Código Transaccion"] = "Código Transacción"

    # Número Lote
    if "Numero Lote" in df.columns and "Número Lote" not in df.columns:
        ren["Numero Lote"] = "Número Lote"

    if "Número lote" in df.columns and "Número Lote" not in df.columns:
        ren["Número lote"] = "Número Lote"

    if "numero lote" in df.columns and "Número Lote" not in df.columns:
        ren["numero lote"] = "Número Lote"

    # Description
    if "Descripcion" in df.columns and "Description" not in df.columns:
        ren["Descripcion"] = "Description"

    if "Descripción" in df.columns and "Description" not in df.columns:
        ren["Descripción"] = "Description"

    if "descripcion" in df.columns and "Description" not in df.columns:
        ren["descripcion"] = "Description"

    if "description" in df.columns and "Description" not in df.columns:
        ren["description"] = "Description"

    # Ubicaciones
    if "Desde Ubicación" in df.columns and "Desde ID Ubicación" not in df.columns:
        ren["Desde Ubicación"] = "Desde ID Ubicación"

    if "A Ubicación" in df.columns and "A ID Ubicación" not in df.columns:
        ren["A Ubicación"] = "A ID Ubicación"

    if "Desde ID Ubicacion" in df.columns and "Desde ID Ubicación" not in df.columns:
        ren["Desde ID Ubicacion"] = "Desde ID Ubicación"

    if "A ID Ubicacion" in df.columns and "A ID Ubicación" not in df.columns:
        ren["A ID Ubicacion"] = "A ID Ubicación"

    if ren:
        df = df.rename(columns=ren)

    return df


# ============================================================
# PARSEO DE FECHA Y HORA
# ============================================================

def _parse_fecha_excel(valor):
    """
    Convierte valores de fecha del Excel WMS a date.
    Formato esperado del WMS: MM/DD/YYYY.
    """
    if valor is None:
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()

    if not texto or texto.lower() == "nan":
        return None

    texto_corto = texto[:10]

    formatos = [
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y"
    ]

    for fmt in formatos:
        try:
            return datetime.strptime(texto_corto, fmt).date()
        except Exception:
            pass

    return None


def _parse_hora_excel(valor):
    """
    Convierte valores de hora del Excel a time.
    Soporta:
    - time
    - datetime
    - timedelta
    - decimal Excel
    - texto HH:MM / HH:MM:SS
    """
    if valor is None:
        return None

    if isinstance(valor, datetime):
        return valor.time().replace(microsecond=0)

    if isinstance(valor, time):
        return valor.replace(microsecond=0)

    if isinstance(valor, timedelta):
        total_segundos = int(valor.total_seconds()) % 86400
        horas = total_segundos // 3600
        minutos = (total_segundos % 3600) // 60
        segundos = total_segundos % 60
        return time(horas, minutos, segundos)

    if isinstance(valor, (int, float)):
        if 0 <= valor < 1:
            total_segundos = int(round(valor * 86400)) % 86400
            horas = total_segundos // 3600
            minutos = (total_segundos % 3600) // 60
            segundos = total_segundos % 60
            return time(horas, minutos, segundos)

    texto = str(valor).strip()

    if not texto or texto.lower() == "nan":
        return None

    formatos = [
        "%H:%M:%S",
        "%H:%M",
        "%I:%M:%S %p",
        "%I:%M %p"
    ]

    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt).time().replace(microsecond=0)
        except Exception:
            pass

    return None


def _build_fechahora(df):
    """
    Construye datetime combinando Fecha + Tiempo.
    """
    if df is None:
        raise ValueError("Error interno: _build_fechahora recibió df=None.")

    return pd.to_datetime(
        df["Fecha"].astype(str).str.strip() + " " + df["Tiempo"].astype(str).str.strip(),
        errors="coerce"
    )


# ============================================================
# NORMALIZACIÓN Y SEGMENTACIÓN
# ============================================================

def _norm_desc_series(s: pd.Series) -> pd.Series:
    """
    Normaliza Description a minúsculas y sin espacios redundantes.
    """
    return (
        s.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(r"\s+", " ", regex=True)
    )


def _segmentar_por_tiempo(df_movs, key_cols, time_col="FechaHora", max_min=15):
    """
    Segmenta dentro de cada llave cuando hay gap > max_min minutos.

    Se usa para consolidar fragmentos de movimientos reales de bodega.
    NO debe usarse para productividad por acciones de despachos.
    """
    if df_movs is None:
        raise ValueError("Error interno: _segmentar_por_tiempo recibió df_movs=None.")

    if df_movs.empty:
        df_movs = df_movs.copy()
        df_movs["Segmento_ID"] = pd.Series(dtype="int64")
        return df_movs

    df_movs = df_movs.sort_values(key_cols + [time_col]).copy()

    prev_t = df_movs.groupby(key_cols, sort=False)[time_col].shift(1)
    gap_min = (df_movs[time_col] - prev_t).dt.total_seconds() / 60

    new_seg = gap_min.isna() | (gap_min > max_min)

    df_movs["Segmento_ID"] = new_seg.groupby(
        [df_movs[c] for c in key_cols]
    ).cumsum()

    return df_movs


def _area_auto_por_pick_norm(pick_norm):
    """
    Clasifica automáticamente el área según el tipo de Pick.
    Se usa cuando una persona está marcada como Mixto.
    """
    pick_norm = str(pick_norm).strip().lower()

    despachos = {
        "loading (pick)",
        "unload/unpick (pick)"
    }

    bodega = {
        "move (pick)",
        "move trailer (pick)",
        "picking (pick)"
    }

    if pick_norm in despachos:
        return "Despachos"

    if pick_norm in bodega:
        return "Bodega"

    return "Bodega"


# ============================================================
# RANKING BODEGA / MOVIMIENTOS
# ============================================================

def _ranking_desde_movimientos(movimientos, area):
    """
    Ranking oficial desde Movimientos Consolidados.

    Bodega se mide por Movimiento_ID real consolidado.
    """
    columnas_salida = [
        "ID Operador",
        "Nombre",
        "Movimientos_Reales"
    ]

    if movimientos is None or movimientos.empty:
        return pd.DataFrame(columns=columnas_salida)

    if "Area_Movimiento" not in movimientos.columns:
        return pd.DataFrame(columns=columnas_salida)

    df_area = movimientos[
        movimientos["Area_Movimiento"]
        .astype(str)
        .str.strip()
        .str.upper()
        .eq(str(area).strip().upper())
    ].copy()

    if df_area.empty:
        return pd.DataFrame(columns=columnas_salida)

    if "Movimiento_ID" not in df_area.columns:
        return pd.DataFrame(columns=columnas_salida)

    if "Nombre" not in df_area.columns:
        return pd.DataFrame(columns=columnas_salida)

    if "ID Operador" not in df_area.columns:
        df_area["ID Operador"] = ""

    df_area["ID Operador"] = (
        df_area["ID Operador"]
        .astype(str)
        .replace("nan", "")
        .replace("<NA>", "")
        .str.strip()
    )

    df_area["Nombre"] = (
        df_area["Nombre"]
        .astype(str)
        .replace("nan", "")
        .replace("<NA>", "")
        .str.strip()
    )

    df_area = df_area[df_area["Nombre"] != ""].copy()

    if df_area.empty:
        return pd.DataFrame(columns=columnas_salida)

    ranking = (
        df_area[["ID Operador", "Nombre", "Movimiento_ID"]]
        .dropna(subset=["Movimiento_ID"])
        .drop_duplicates()
        .groupby(["ID Operador", "Nombre"], dropna=False)["Movimiento_ID"]
        .count()
        .reset_index(name="Movimientos_Reales")
        .sort_values("Movimientos_Reales", ascending=False)
        .reset_index(drop=True)
    )

    if ranking.empty:
        return pd.DataFrame(columns=columnas_salida)

    total = int(ranking["Movimientos_Reales"].sum())

    fila_total = pd.DataFrame([{
        "ID Operador": "",
        "Nombre": "TOTAL",
        "Movimientos_Reales": total
    }])

    ranking = pd.concat([ranking, fila_total], ignore_index=True)

    return ranking[columnas_salida]


def _ranking_bodega_desde_movimientos(movimientos):
    """
    Alias para Ranking Bodega.
    """
    return _ranking_desde_movimientos(movimientos, "Bodega")


# ============================================================
# DESPACHOS - EVENTOS POR FILA
# ============================================================

def _puerta_relacion_loading_pick(ubicacion_desde):
    """
    Para Loading (pick):
    Desde ID Ubicación puede ser ZFDOOR3-23A-C.
    La puerta de relación para el vehículo es ZFDOOR3-23A.

    Importante:
    No se modifica la ubicación original.
    Esta función solo crea una llave auxiliar para relacionar ciclo.
    """
    texto = str(ubicacion_desde).strip()

    if texto.endswith("-C"):
        return texto[:-2]

    return texto


def _puerta_relacion_loading_put(ubicacion_a):
    """
    Para Loading (put):
    A ID Ubicación representa la puerta/vehículo sin -C.
    """
    return str(ubicacion_a).strip()


def _construir_eventos_despacho_por_fila(df_base, asignacion_operadores_area=None, nombres_ignorar=None):
    """
    Construye eventos de despacho por fila, sin consolidar.

    Reglas:
    - Loading (pick) cuenta como acción real de toma/cargue.
    - Loading (put) cuenta como acción real de soltado en vehículo.
    - Unload/Unpick (pick) cuenta como acción real de toma para devolución.
    - Unload/Unpick (put) puede ser espejo del sistema o put final real.
    - LP NO se usa como llave de relación.
    """

    columnas_eventos = [
        "Fila_Origen",
        "ID Operador",
        "Nombre_norm",
        "Nombre",
        "Description_norm",
        "Item Number",
        "Número Lote",
        "LP",
        "Desde ID Ubicación",
        "A ID Ubicación",
        "FechaHora",
        "Cantidad",
        "Area_Operador_Asignada",
        "Tipo_Evento_Despacho",
        "Tipo_Flujo_Despacho",
        "Cuenta_Productividad_Despacho",
        "Puerta_Relacion",
        "Evento_Despacho_ID"
    ]

    if df_base is None or df_base.empty:
        return pd.DataFrame(columns=columnas_eventos)

    df = df_base.copy()

    df = _ensure_columns(df, [
        "ID Operador",
        "Nombre_norm",
        "Nombre",
        "Description_norm",
        "Item Number",
        "Número Lote",
        "LP",
        "Desde ID Ubicación",
        "A ID Ubicación",
        "FechaHora",
        "Cantidad"
    ])

    if asignacion_operadores_area is None:
        asignacion_operadores_area = {}

    if nombres_ignorar is None:
        nombres_ignorar = set()

    df["Area_Operador_Asignada"] = (
        df["Nombre_norm"]
        .astype(str)
        .map(asignacion_operadores_area)
        .fillna("Excluir")
    )

    descs_despacho = {
        "loading (pick)",
        "loading (put)",
        "unload/unpick (pick)",
        "unload/unpick (put)"
    }

    df = df[df["Description_norm"].isin(descs_despacho)].copy()
    df = df[~df["Nombre_norm"].astype(str).isin(nombres_ignorar)].copy()
    df = df[df["Area_Operador_Asignada"].isin(["Despachos", "Mixto"])].copy()

    if df.empty:
        return pd.DataFrame(columns=columnas_eventos)

    df = df.reset_index(drop=False).rename(columns={"index": "Fila_Origen"})

    eventos = df[[
        "Fila_Origen",
        "ID Operador",
        "Nombre_norm",
        "Nombre",
        "Description_norm",
        "Item Number",
        "Número Lote",
        "LP",
        "Desde ID Ubicación",
        "A ID Ubicación",
        "FechaHora",
        "Cantidad",
        "Area_Operador_Asignada"
    ]].copy()

    eventos["Tipo_Evento_Despacho"] = ""
    eventos["Tipo_Flujo_Despacho"] = ""
    eventos["Cuenta_Productividad_Despacho"] = False
    eventos["Puerta_Relacion"] = ""

    # Loading pick
    mask_loading_pick = eventos["Description_norm"] == "loading (pick)"
    eventos.loc[mask_loading_pick, "Tipo_Evento_Despacho"] = "CARGUE_PICK_REAL"
    eventos.loc[mask_loading_pick, "Tipo_Flujo_Despacho"] = "CARGUE_VEHICULO"
    eventos.loc[mask_loading_pick, "Cuenta_Productividad_Despacho"] = True
    eventos.loc[mask_loading_pick, "Puerta_Relacion"] = eventos.loc[
        mask_loading_pick, "Desde ID Ubicación"
    ].apply(_puerta_relacion_loading_pick)

    # Loading put
    mask_loading_put = eventos["Description_norm"] == "loading (put)"
    eventos.loc[mask_loading_put, "Tipo_Evento_Despacho"] = "CARGUE_PUT_REAL"
    eventos.loc[mask_loading_put, "Tipo_Flujo_Despacho"] = "CARGUE_VEHICULO"
    eventos.loc[mask_loading_put, "Cuenta_Productividad_Despacho"] = True
    eventos.loc[mask_loading_put, "Puerta_Relacion"] = eventos.loc[
        mask_loading_put, "A ID Ubicación"
    ].apply(_puerta_relacion_loading_put)

    # Devolución pick real
    mask_unload_pick = eventos["Description_norm"] == "unload/unpick (pick)"
    eventos.loc[mask_unload_pick, "Tipo_Evento_Despacho"] = "DEVOLUCION_PICK_REAL"
    eventos.loc[mask_unload_pick, "Tipo_Flujo_Despacho"] = "DEVOLUCION_NO_CONFORME"
    eventos.loc[mask_unload_pick, "Cuenta_Productividad_Despacho"] = True
    eventos.loc[mask_unload_pick, "Puerta_Relacion"] = eventos.loc[
        mask_unload_pick, "Desde ID Ubicación"
    ].astype(str).str.strip()

    # Devolución put: primero marcar como final real, luego detectar espejos
    mask_unload_put = eventos["Description_norm"] == "unload/unpick (put)"
    eventos.loc[mask_unload_put, "Tipo_Evento_Despacho"] = "DEVOLUCION_PUT_FINAL_REAL"
    eventos.loc[mask_unload_put, "Tipo_Flujo_Despacho"] = "DEVOLUCION_NO_CONFORME"
    eventos.loc[mask_unload_put, "Cuenta_Productividad_Despacho"] = True
    eventos.loc[mask_unload_put, "Puerta_Relacion"] = eventos.loc[
        mask_unload_put, "Desde ID Ubicación"
    ].astype(str).str.strip()

    eventos = _marcar_puts_espejo_devolucion(eventos)

    eventos["Evento_Despacho_ID"] = (
        "DSP-" + (eventos.index + 1).astype(str).str.zfill(7)
    )

    return eventos[columnas_eventos]


def _marcar_puts_espejo_devolucion(eventos):
    """
    Marca Unload/Unpick (put) espejo del sistema.

    Regla:
    El put espejo tiene mismas ubicaciones que el Unload/Unpick (pick)
    del cual es espejo.

    Se usa llave conservadora:
    Nombre_norm + Item Number + Número Lote + Desde ID Ubicación
    + A ID Ubicación + Cantidad + FechaHora cercana.
    """
    if eventos is None or eventos.empty:
        return eventos

    df = eventos.copy()

    picks = df[df["Description_norm"] == "unload/unpick (pick)"].copy()
    puts = df[df["Description_norm"] == "unload/unpick (put)"].copy()

    if picks.empty or puts.empty:
        return df

    picks["_key_espejo"] = (
        picks["Nombre_norm"].astype(str).str.strip() + "|" +
        picks["Item Number"].astype(str).str.strip() + "|" +
        picks["Número Lote"].astype(str).str.strip() + "|" +
        picks["Desde ID Ubicación"].astype(str).str.strip() + "|" +
        picks["A ID Ubicación"].astype(str).str.strip() + "|" +
        picks["Cantidad"].astype(str).str.strip()
    )

    puts["_key_espejo"] = (
        puts["Nombre_norm"].astype(str).str.strip() + "|" +
        puts["Item Number"].astype(str).str.strip() + "|" +
        puts["Número Lote"].astype(str).str.strip() + "|" +
        puts["Desde ID Ubicación"].astype(str).str.strip() + "|" +
        puts["A ID Ubicación"].astype(str).str.strip() + "|" +
        puts["Cantidad"].astype(str).str.strip()
    )

    keys_pick = set(picks["_key_espejo"].dropna().astype(str).tolist())

    mask_put_espejo = puts["_key_espejo"].astype(str).isin(keys_pick)

    indices_espejo = puts.loc[mask_put_espejo].index

    df.loc[indices_espejo, "Tipo_Evento_Despacho"] = "DEVOLUCION_PUT_ESPEJO_SISTEMA"
    df.loc[indices_espejo, "Cuenta_Productividad_Despacho"] = False

    return df


# ============================================================
# DESPACHOS - CICLOS PRODUCTIVOS
# ============================================================

def _construir_ciclos_despacho(eventos_despacho, max_gap_min=60):
    """
    Construye ciclos productivos de despacho.

    Cargue:
    - Loading pick(s) + Loading put(s)
    - Agrupa por Nombre_norm + Puerta_Relacion + ventana temporal.

    Devolución:
    - Unload/Unpick pick real(s) + put final real
    - Excluye puts espejo del sistema.
    - Agrupa por Nombre_norm + Item Number + Número Lote + ventana temporal.
    """
    columnas = [
        "Ciclo_Despacho_ID",
        "Nombre_norm",
        "ID Operador",
        "Nombre",
        "Tipo_Flujo_Despacho",
        "Puerta_Relacion",
        "Item Number",
        "Número Lote",
        "FH_inicio",
        "FH_fin",
        "Tiempo_Duracion",
        "Acciones_Pick",
        "Acciones_Put",
        "Acciones_Productivas",
        "Cantidad_Pick_Total",
        "Cantidad_Put_Total",
        "Referencias_Distintas",
        "Lotes_Distintos",
        "Estado_Ciclo",
        "Estado_Cantidad"
    ]

    if eventos_despacho is None or eventos_despacho.empty:
        return pd.DataFrame(columns=columnas)

    eventos = eventos_despacho.copy()
    eventos["FechaHora"] = pd.to_datetime(eventos["FechaHora"], errors="coerce")
    eventos = eventos[eventos["FechaHora"].notna()].copy()

    if eventos.empty:
        return pd.DataFrame(columns=columnas)

    ciclos = []

    # -------------------------------
    # Ciclos de cargue
    # -------------------------------
    cargue = eventos[
        eventos["Tipo_Flujo_Despacho"].eq("CARGUE_VEHICULO")
    ].copy()

    if not cargue.empty:
        cargue = cargue.sort_values(["Nombre_norm", "Puerta_Relacion", "FechaHora"]).copy()

        group_cols = ["Nombre_norm", "Puerta_Relacion"]

        cargue["Gap_Min"] = (
            cargue.groupby(group_cols, sort=False)["FechaHora"]
            .diff()
            .dt.total_seconds()
            .div(60)
        )

        cargue["Nuevo_Ciclo"] = cargue["Gap_Min"].isna() | (cargue["Gap_Min"] > max_gap_min)

        cargue["Grupo_Ciclo"] = (
            cargue.groupby(group_cols, sort=False)["Nuevo_Ciclo"]
            .cumsum()
        )

        for _, g in cargue.groupby(group_cols + ["Grupo_Ciclo"], sort=False):
            acciones_pick = int((g["Tipo_Evento_Despacho"] == "CARGUE_PICK_REAL").sum())
            acciones_put = int((g["Tipo_Evento_Despacho"] == "CARGUE_PUT_REAL").sum())

            cantidad_pick = pd.to_numeric(
                g.loc[g["Tipo_Evento_Despacho"] == "CARGUE_PICK_REAL", "Cantidad"],
                errors="coerce"
            ).fillna(0).sum()

            cantidad_put = pd.to_numeric(
                g.loc[g["Tipo_Evento_Despacho"] == "CARGUE_PUT_REAL", "Cantidad"],
                errors="coerce"
            ).fillna(0).sum()

            estado = "CICLO_CARGUE_COMPLETO"
            if acciones_pick > 0 and acciones_put == 0:
                estado = "CARGUE_PICK_SIN_PUT"
            elif acciones_pick == 0 and acciones_put > 0:
                estado = "CARGUE_PUT_SIN_PICK"

            estado_cantidad = "NO_APLICA"

            ciclos.append({
                "Nombre_norm": g["Nombre_norm"].iloc[0],
                "ID Operador": g["ID Operador"].iloc[0] if "ID Operador" in g.columns else "",
                "Nombre": g["Nombre"].iloc[0],
                "Tipo_Flujo_Despacho": "CARGUE_VEHICULO",
                "Puerta_Relacion": g["Puerta_Relacion"].iloc[0],
                "Item Number": "",
                "Número Lote": "",
                "FH_inicio": g["FechaHora"].min(),
                "FH_fin": g["FechaHora"].max(),
                "Acciones_Pick": acciones_pick,
                "Acciones_Put": acciones_put,
                "Acciones_Productivas": int(g["Cuenta_Productividad_Despacho"].sum()),
                "Cantidad_Pick_Total": cantidad_pick,
                "Cantidad_Put_Total": cantidad_put,
                "Referencias_Distintas": int(g["Item Number"].nunique()),
                "Lotes_Distintos": int(g["Número Lote"].nunique()),
                "Estado_Ciclo": estado,
                "Estado_Cantidad": estado_cantidad
            })

    # -------------------------------
    # Ciclos de devolución
    # -------------------------------
    dev = eventos[
        eventos["Tipo_Flujo_Despacho"].eq("DEVOLUCION_NO_CONFORME")
    ].copy()

    if not dev.empty:
        dev = dev[
            dev["Tipo_Evento_Despacho"].isin([
                "DEVOLUCION_PICK_REAL",
                "DEVOLUCION_PUT_FINAL_REAL"
            ])
        ].copy()

        if not dev.empty:
            dev = dev.sort_values([
                "Nombre_norm",
                "Item Number",
                "Número Lote",
                "FechaHora"
            ]).copy()

            group_cols = ["Nombre_norm", "Item Number", "Número Lote"]

            dev["Gap_Min"] = (
                dev.groupby(group_cols, sort=False)["FechaHora"]
                .diff()
                .dt.total_seconds()
                .div(60)
            )

            dev["Nuevo_Ciclo"] = dev["Gap_Min"].isna() | (dev["Gap_Min"] > max_gap_min)

            dev["Grupo_Ciclo"] = (
                dev.groupby(group_cols, sort=False)["Nuevo_Ciclo"]
                .cumsum()
            )

            for _, g in dev.groupby(group_cols + ["Grupo_Ciclo"], sort=False):
                acciones_pick = int((g["Tipo_Evento_Despacho"] == "DEVOLUCION_PICK_REAL").sum())
                acciones_put = int((g["Tipo_Evento_Despacho"] == "DEVOLUCION_PUT_FINAL_REAL").sum())

                cantidad_pick = pd.to_numeric(
                    g.loc[g["Tipo_Evento_Despacho"] == "DEVOLUCION_PICK_REAL", "Cantidad"],
                    errors="coerce"
                ).fillna(0).sum()

                cantidad_put = pd.to_numeric(
                    g.loc[g["Tipo_Evento_Despacho"] == "DEVOLUCION_PUT_FINAL_REAL", "Cantidad"],
                    errors="coerce"
                ).fillna(0).sum()

                estado = "CICLO_DEVOLUCION_COMPLETO"
                if acciones_pick > 0 and acciones_put == 0:
                    estado = "DEVOLUCION_PICK_SIN_PUT_FINAL"
                elif acciones_pick == 0 and acciones_put > 0:
                    estado = "DEVOLUCION_PUT_FINAL_SIN_PICK"

                estado_cantidad = "OK"
                if acciones_pick > 0 and acciones_put > 0:
                    if round(float(cantidad_pick), 6) != round(float(cantidad_put), 6):
                        estado_cantidad = "REVISAR_DIFERENCIA_CANTIDAD"
                else:
                    estado_cantidad = "NO_APLICA"

                ciclos.append({
                    "Nombre_norm": g["Nombre_norm"].iloc[0],
                    "ID Operador": g["ID Operador"].iloc[0] if "ID Operador" in g.columns else "",
                    "Nombre": g["Nombre"].iloc[0],
                    "Tipo_Flujo_Despacho": "DEVOLUCION_NO_CONFORME",
                    "Puerta_Relacion": g["Puerta_Relacion"].iloc[0] if "Puerta_Relacion" in g.columns else "",
                    "Item Number": g["Item Number"].iloc[0],
                    "Número Lote": g["Número Lote"].iloc[0],
                    "FH_inicio": g["FechaHora"].min(),
                    "FH_fin": g["FechaHora"].max(),
                    "Acciones_Pick": acciones_pick,
                    "Acciones_Put": acciones_put,
                    "Acciones_Productivas": int(g["Cuenta_Productividad_Despacho"].sum()),
                    "Cantidad_Pick_Total": cantidad_pick,
                    "Cantidad_Put_Total": cantidad_put,
                    "Referencias_Distintas": int(g["Item Number"].nunique()),
                    "Lotes_Distintos": int(g["Número Lote"].nunique()),
                    "Estado_Ciclo": estado,
                    "Estado_Cantidad": estado_cantidad
                })

    if not ciclos:
        return pd.DataFrame(columns=columnas)

    df_ciclos = pd.DataFrame(ciclos)

    df_ciclos["FH_inicio"] = pd.to_datetime(df_ciclos["FH_inicio"], errors="coerce")
    df_ciclos["FH_fin"] = pd.to_datetime(df_ciclos["FH_fin"], errors="coerce")

    df_ciclos["Tiempo_Duracion"] = (
        (df_ciclos["FH_fin"] - df_ciclos["FH_inicio"])
        .dt.total_seconds()
        .div(60)
        .round(2)
    )

    df_ciclos = df_ciclos.sort_values(["Nombre", "FH_inicio"]).reset_index(drop=True)

    df_ciclos["Ciclo_Despacho_ID"] = (
        "CDSP-" + (df_ciclos.index + 1).astype(str).str.zfill(6)
    )

    return df_ciclos[columnas]


def _calcular_inactividad_despacho(ciclos_despacho, fecha_str, ini, fin, min_inactividad_reportable=5):
    """
    Calcula lapsos de inactividad entre ciclos productivos de Despacho.

    Salida:
    - Incluye ID Operador.
    - No incluye Nombre_norm.
    """

    columnas = [
        "ID Operador",
        "Nombre",
        "Tipo_Inactividad",
        "Fin_Actividad_Anterior",
        "Inicio_Actividad_Siguiente",
        "Tiempo_Inactivo",
        "Fecha"
    ]

    if ciclos_despacho is None or ciclos_despacho.empty:
        return pd.DataFrame(columns=columnas)

    ciclos = ciclos_despacho.copy()

    ciclos = _ensure_columns(ciclos, [
        "ID Operador",
        "Nombre_norm",
        "Nombre",
        "FH_inicio",
        "FH_fin",
        "Estado_Ciclo"
    ])

    ciclos["FH_inicio"] = pd.to_datetime(ciclos["FH_inicio"], errors="coerce")
    ciclos["FH_fin"] = pd.to_datetime(ciclos["FH_fin"], errors="coerce")

    ciclos = ciclos[
        ciclos["FH_inicio"].notna() &
        ciclos["FH_fin"].notna() &
        ciclos["Estado_Ciclo"].astype(str).str.contains("COMPLETO", na=False)
    ].copy()

    if ciclos.empty:
        return pd.DataFrame(columns=columnas)

    fecha_base = datetime.strptime(fecha_str.strip(), "%d/%m/%Y").date()
    h_ini = datetime.strptime(ini, "%H:%M").time()
    h_fin = datetime.strptime(fin, "%H:%M").time()

    if h_ini <= h_fin:
        dt_ini_turno = datetime.combine(fecha_base, h_ini)
        dt_fin_turno = datetime.combine(fecha_base, h_fin)
    else:
        dt_ini_turno = datetime.combine(fecha_base - timedelta(days=1), h_ini)
        dt_fin_turno = datetime.combine(fecha_base, h_fin)

    registros = []

    group_cols = ["ID Operador", "Nombre_norm", "Nombre"]

    for _, g in ciclos.groupby(group_cols, dropna=False):
        g = g.sort_values("FH_inicio").copy()

        id_operador = str(g["ID Operador"].iloc[0]).strip()
        nombre = str(g["Nombre"].iloc[0]).strip()

        primera = g.iloc[0]

        # ========================================================
        # 1. Inicio turno a primera actividad
        # ========================================================
        gap_ini = (primera["FH_inicio"] - dt_ini_turno).total_seconds() / 60

        if gap_ini >= min_inactividad_reportable:
            registros.append({
                "ID Operador": id_operador,
                "Nombre": nombre,
                "Tipo_Inactividad": "INICIO_TURNO_A_PRIMERA_ACTIVIDAD",
                "Fin_Actividad_Anterior": dt_ini_turno,
                "Inicio_Actividad_Siguiente": primera["FH_inicio"],
                "Tiempo_Inactivo": round(gap_ini, 2),
                "Fecha": fecha_str
            })

        # ========================================================
        # 2. Entre ciclos
        # ========================================================
        prev_fin = primera["FH_fin"]

        for i in range(1, len(g)):
            row = g.iloc[i]

            gap = (row["FH_inicio"] - prev_fin).total_seconds() / 60

            if gap >= min_inactividad_reportable:
                registros.append({
                    "ID Operador": id_operador,
                    "Nombre": nombre,
                    "Tipo_Inactividad": "ENTRE_CICLOS",
                    "Fin_Actividad_Anterior": prev_fin,
                    "Inicio_Actividad_Siguiente": row["FH_inicio"],
                    "Tiempo_Inactivo": round(gap, 2),
                    "Fecha": fecha_str
                })

            if row["FH_fin"] > prev_fin:
                prev_fin = row["FH_fin"]

        # ========================================================
        # 3. Última actividad a fin turno
        # ========================================================
        gap_fin = (dt_fin_turno - prev_fin).total_seconds() / 60

        if gap_fin >= min_inactividad_reportable:
            registros.append({
                "ID Operador": id_operador,
                "Nombre": nombre,
                "Tipo_Inactividad": "ULTIMA_ACTIVIDAD_A_FIN_TURNO",
                "Fin_Actividad_Anterior": prev_fin,
                "Inicio_Actividad_Siguiente": dt_fin_turno,
                "Tiempo_Inactivo": round(gap_fin, 2),
                "Fecha": fecha_str
            })

    if not registros:
        return pd.DataFrame(columns=columnas)

    resultado = pd.DataFrame(registros)

    resultado = resultado.sort_values([
        "Nombre",
        "Fin_Actividad_Anterior",
        "Inicio_Actividad_Siguiente"
    ]).reset_index(drop=True)

    return resultado[columnas]


# ============================================================
# RANKING DESPACHOS
# ============================================================

def _ranking_despachos_desde_eventos(eventos_despacho):
    """
    Ranking básico de Despachos por eventos productivos.
    Se conserva por compatibilidad.

    Cuenta:
    - CARGUE_PICK_REAL
    - CARGUE_PUT_REAL
    - DEVOLUCION_PICK_REAL
    - DEVOLUCION_PUT_FINAL_REAL

    No cuenta:
    - DEVOLUCION_PUT_ESPEJO_SISTEMA
    """
    columnas_salida = [
        "ID Operador",
        "Nombre",
        "Cargues_Vehiculo",
        "Devoluciones",
        "Total_Eventos_Despacho"
    ]

    if eventos_despacho is None or eventos_despacho.empty:
        return pd.DataFrame(columns=columnas_salida)

    df = eventos_despacho.copy()

    for col in ["ID Operador", "Nombre", "Tipo_Evento_Despacho", "Evento_Despacho_ID"]:
        if col not in df.columns:
            df[col] = pd.NA

    if "Cuenta_Productividad_Despacho" not in df.columns:
        df["Cuenta_Productividad_Despacho"] = True

    df["ID Operador"] = (
        df["ID Operador"]
        .astype(str)
        .replace("nan", "")
        .replace("<NA>", "")
        .str.strip()
    )

    df["Nombre"] = (
        df["Nombre"]
        .astype(str)
        .replace("nan", "")
        .replace("<NA>", "")
        .str.strip()
    )

    df["Tipo_Evento_Despacho"] = (
        df["Tipo_Evento_Despacho"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    df = df[df["Nombre"].astype(str).str.strip() != ""].copy()

    if df.empty:
        return pd.DataFrame(columns=columnas_salida)

    df_prod = df[df["Cuenta_Productividad_Despacho"] == True].copy()

    cargues = (
        df_prod[df_prod["Tipo_Evento_Despacho"].isin(["CARGUE_PICK_REAL", "CARGUE_PUT_REAL"])]
        .groupby(["ID Operador", "Nombre"], dropna=False)
        .size()
        .reset_index(name="Cargues_Vehiculo")
    )

    devoluciones = (
        df_prod[df_prod["Tipo_Evento_Despacho"].isin(["DEVOLUCION_PICK_REAL", "DEVOLUCION_PUT_FINAL_REAL"])]
        .groupby(["ID Operador", "Nombre"], dropna=False)
        .size()
        .reset_index(name="Devoluciones")
    )

    ranking = cargues.merge(
        devoluciones,
        on=["ID Operador", "Nombre"],
        how="outer"
    ).fillna(0)

    if ranking.empty:
        return pd.DataFrame(columns=columnas_salida)

    ranking["Cargues_Vehiculo"] = ranking["Cargues_Vehiculo"].astype(int)
    ranking["Devoluciones"] = ranking["Devoluciones"].astype(int)

    ranking["Total_Eventos_Despacho"] = (
        ranking["Cargues_Vehiculo"] + ranking["Devoluciones"]
    )

    ranking = ranking.sort_values(
        "Total_Eventos_Despacho",
        ascending=False
    ).reset_index(drop=True)

    fila_total = pd.DataFrame([{
        "ID Operador": "",
        "Nombre": "TOTAL",
        "Cargues_Vehiculo": int(ranking["Cargues_Vehiculo"].sum()),
        "Devoluciones": int(ranking["Devoluciones"].sum()),
        "Total_Eventos_Despacho": int(ranking["Total_Eventos_Despacho"].sum())
    }])

    ranking = pd.concat([ranking, fila_total], ignore_index=True)

    return ranking[columnas_salida]


def _ranking_despachos_desde_ciclos(eventos_despacho, ciclos_despacho, inactividad_despacho=None):
    """
    Ranking avanzado de Despachos con acciones, ciclos, tiempos e inactividad.
    """
    columnas = [
        "ID Operador",
        "Nombre",
        "Acciones_Loading_Pick",
        "Acciones_Loading_Put",
        "Acciones_Unload_Pick",
        "Acciones_Unload_Put_Final",
        "Puts_Unload_Espejo_Sistema",
        "Total_Acciones_Productivas",
        "Ciclos_Cargue_Completados",
        "Ciclos_Devolucion_Completados",
        "Minutos_Productivos",
        "Tiempo_Inactivo",
        "Porcentaje_Utilizacion",
        "Duracion_Promedio_Cargue",
        "Duracion_Promedio_Devolucion"
    ]

    if eventos_despacho is None or eventos_despacho.empty:
        return pd.DataFrame(columns=columnas)

    eventos = eventos_despacho.copy()

    eventos = _ensure_columns(eventos, [
        "ID Operador",
        "Nombre",
        "Tipo_Evento_Despacho",
        "Cuenta_Productividad_Despacho"
    ])

    eventos["ID Operador"] = (
        eventos["ID Operador"].astype(str).replace("nan", "").replace("<NA>", "").str.strip()
    )

    eventos["Nombre"] = (
        eventos["Nombre"].astype(str).replace("nan", "").replace("<NA>", "").str.strip()
    )

    eventos = eventos[eventos["Nombre"] != ""].copy()

    if eventos.empty:
        return pd.DataFrame(columns=columnas)

    base = (
        eventos[["ID Operador", "Nombre"]]
        .drop_duplicates()
        .reset_index(drop=True)
    )

    def conteo_evento(tipo):
        return (
            eventos[eventos["Tipo_Evento_Despacho"] == tipo]
            .groupby(["ID Operador", "Nombre"], dropna=False)
            .size()
            .reset_index(name=tipo)
        )

    counts = base.copy()

    mapping = {
        "CARGUE_PICK_REAL": "Acciones_Loading_Pick",
        "CARGUE_PUT_REAL": "Acciones_Loading_Put",
        "DEVOLUCION_PICK_REAL": "Acciones_Unload_Pick",
        "DEVOLUCION_PUT_FINAL_REAL": "Acciones_Unload_Put_Final",
        "DEVOLUCION_PUT_ESPEJO_SISTEMA": "Puts_Unload_Espejo_Sistema"
    }

    for tipo, col_salida in mapping.items():
        tmp = conteo_evento(tipo).rename(columns={tipo: col_salida})
        counts = counts.merge(tmp, on=["ID Operador", "Nombre"], how="left")

    for col in mapping.values():
        counts[col] = counts[col].fillna(0).astype(int)

    counts["Total_Acciones_Productivas"] = (
        counts["Acciones_Loading_Pick"] +
        counts["Acciones_Loading_Put"] +
        counts["Acciones_Unload_Pick"] +
        counts["Acciones_Unload_Put_Final"]
    )

    if ciclos_despacho is None or ciclos_despacho.empty:
        counts["Ciclos_Cargue_Completados"] = 0
        counts["Ciclos_Devolucion_Completados"] = 0
        counts["Minutos_Productivos"] = 0.0
        counts["Duracion_Promedio_Cargue"] = 0.0
        counts["Duracion_Promedio_Devolucion"] = 0.0
    else:
        ciclos = ciclos_despacho.copy()
        ciclos = _ensure_columns(ciclos, [
            "ID Operador",
            "Nombre",
            "Tipo_Flujo_Despacho",
            "Estado_Ciclo",
            "Tiempo_Duracion"
        ])

        ciclos["ID Operador"] = ciclos["ID Operador"].astype(str).replace("nan", "").replace("<NA>", "").str.strip()
        ciclos["Nombre"] = ciclos["Nombre"].astype(str).replace("nan", "").replace("<NA>", "").str.strip()

        ciclos_completos = ciclos[
            ciclos["Estado_Ciclo"].astype(str).str.contains("COMPLETO", na=False)
        ].copy()

        resumen_ciclos = (
            ciclos_completos
            .groupby(["ID Operador", "Nombre"], dropna=False)
            .agg(
                Ciclos_Cargue_Completados=(
                    "Tipo_Flujo_Despacho",
                    lambda s: int((s == "CARGUE_VEHICULO").sum())
                ),
                Ciclos_Devolucion_Completados=(
                    "Tipo_Flujo_Despacho",
                    lambda s: int((s == "DEVOLUCION_NO_CONFORME").sum())
                ),
                Minutos_Productivos=("Tiempo_Duracion", "sum")
            )
            .reset_index()
        )

        prom_cargue = (
            ciclos_completos[ciclos_completos["Tipo_Flujo_Despacho"] == "CARGUE_VEHICULO"]
            .groupby(["ID Operador", "Nombre"], dropna=False)["Tiempo_Duracion"]
            .mean()
            .reset_index(name="Duracion_Promedio_Cargue")
        )

        prom_dev = (
            ciclos_completos[ciclos_completos["Tipo_Flujo_Despacho"] == "DEVOLUCION_NO_CONFORME"]
            .groupby(["ID Operador", "Nombre"], dropna=False)["Tiempo_Duracion"]
            .mean()
            .reset_index(name="Duracion_Promedio_Devolucion")
        )

        counts = counts.merge(resumen_ciclos, on=["ID Operador", "Nombre"], how="left")
        counts = counts.merge(prom_cargue, on=["ID Operador", "Nombre"], how="left")
        counts = counts.merge(prom_dev, on=["ID Operador", "Nombre"], how="left")

        for col in [
            "Ciclos_Cargue_Completados",
            "Ciclos_Devolucion_Completados",
            "Minutos_Productivos",
            "Duracion_Promedio_Cargue",
            "Duracion_Promedio_Devolucion"
        ]:
            counts[col] = counts[col].fillna(0)

    if inactividad_despacho is None or inactividad_despacho.empty:
        counts["Tiempo_Inactivo"] = 0.0
    else:
        ina = inactividad_despacho.copy()
        ina = _ensure_columns(ina, ["Nombre", "Tiempo_Inactivo"])

        resumen_ina = (
            ina.groupby("Nombre", dropna=False)["Tiempo_Inactivo"]
            .sum()
            .reset_index(name="Tiempo_Inactivo")
        )

        counts = counts.merge(resumen_ina, on="Nombre", how="left")
        counts["Tiempo_Inactivo"] = counts["Tiempo_Inactivo"].fillna(0)

    total_tiempo = counts["Minutos_Productivos"] + counts["Tiempo_Inactivo"]

    counts["Porcentaje_Utilizacion"] = 0.0
    mask_total = total_tiempo > 0
    counts.loc[mask_total, "Porcentaje_Utilizacion"] = (
        counts.loc[mask_total, "Minutos_Productivos"] / total_tiempo.loc[mask_total]
    )

    for col in [
        "Minutos_Productivos",
        "Tiempo_Inactivo",
        "Duracion_Promedio_Cargue",
        "Duracion_Promedio_Devolucion",
        "Porcentaje_Utilizacion"
    ]:
        counts[col] = counts[col].astype(float).round(4)

    counts = counts.sort_values(
        "Total_Acciones_Productivas",
        ascending=False
    ).reset_index(drop=True)

    if not counts.empty:
        fila_total = {
            "ID Operador": "",
            "Nombre": "TOTAL"
        }

        for col in columnas:
            if col in ["ID Operador", "Nombre"]:
                continue

            if col == "Porcentaje_Utilizacion":
                prod = counts["Minutos_Productivos"].sum()
                ina = counts["Tiempo_Inactivo"].sum()
                fila_total[col] = round(prod / (prod + ina), 4) if (prod + ina) > 0 else 0.0
            elif col.startswith("Duracion_Promedio"):
                fila_total[col] = ""
            else:
                fila_total[col] = counts[col].sum()

        counts = pd.concat([counts, pd.DataFrame([fila_total])], ignore_index=True)

    return counts[columnas]


# ============================================================
# ANÁLISIS HORARIO
# ============================================================

def _generar_rangos_horarios_turno(fecha_str, ini, fin):
    """
    Genera rangos horarios dinámicos según el turno seleccionado.

    Retorna lista de dicts:
    [
        {
            "label": "7:30 a 8:30",
            "inicio": datetime,
            "fin": datetime
        }
    ]
    """
    fecha_base = datetime.strptime(fecha_str.strip(), "%d/%m/%Y").date()
    h_ini = datetime.strptime(ini, "%H:%M").time()
    h_fin = datetime.strptime(fin, "%H:%M").time()

    if h_ini <= h_fin:
        dt_ini = datetime.combine(fecha_base, h_ini)
        dt_fin = datetime.combine(fecha_base, h_fin)
    else:
        dt_ini = datetime.combine(fecha_base - timedelta(days=1), h_ini)
        dt_fin = datetime.combine(fecha_base, h_fin)

    rangos = []
    actual = dt_ini

    while actual < dt_fin:
        siguiente = actual + timedelta(hours=1)

        if siguiente > dt_fin:
            siguiente = dt_fin

        label = f"{actual.strftime('%H:%M')} a {siguiente.strftime('%H:%M')}"

        rangos.append({
            "label": label,
            "inicio": actual,
            "fin": siguiente
        })

        actual = siguiente

    return rangos


def _asignar_label_rango_horario(fecha_hora, rangos):
    """
    Devuelve el label del rango horario al que pertenece fecha_hora.
    """
    if pd.isna(fecha_hora):
        return None

    fh = pd.to_datetime(fecha_hora, errors="coerce")

    if pd.isna(fh):
        return None

    for r in rangos:
        if r["inicio"] <= fh < r["fin"]:
            return r["label"]

    if rangos and fh == rangos[-1]["fin"]:
        return rangos[-1]["label"]

    return None


def _generar_matriz_horaria(
    df_base,
    fecha_str,
    ini,
    fin,
    meta_ppt,
    valor_col,
    aggfunc="sum",
    total_col="Total"
):
    """
    Genera matriz horaria estándar.

    Para Despachos:
        NOMBRES | Horas | Total | Meta PPT | % Productividad | Fecha

    Para Bodega:
        NOMBRES | Horas | Total Movimientos | Meta PPT | % Productividad | Fecha
    """

    rangos = _generar_rangos_horarios_turno(fecha_str, ini, fin)
    labels = [r["label"] for r in rangos]

    pct_col = "% Productividad"
    fecha_col = "Fecha"

    columnas_salida = ["NOMBRES"] + labels + [
        total_col,
        "Meta PPT",
        pct_col,
        fecha_col
    ]

    if df_base is None or df_base.empty:
        return pd.DataFrame(columns=columnas_salida)

    df = df_base.copy()

    df = _ensure_columns(df, ["Nombre", "FechaHora_Analisis", valor_col])

    df["Nombre"] = (
        df["Nombre"]
        .astype(str)
        .replace("nan", "")
        .replace("<NA>", "")
        .str.strip()
    )

    df = df[df["Nombre"] != ""].copy()

    if df.empty:
        return pd.DataFrame(columns=columnas_salida)

    df["FechaHora_Analisis"] = pd.to_datetime(
        df["FechaHora_Analisis"],
        errors="coerce"
    )

    df = df[df["FechaHora_Analisis"].notna()].copy()

    if df.empty:
        return pd.DataFrame(columns=columnas_salida)

    df["Rango_Hora"] = df["FechaHora_Analisis"].apply(
        lambda x: _asignar_label_rango_horario(x, rangos)
    )

    df = df[df["Rango_Hora"].notna()].copy()

    if df.empty:
        return pd.DataFrame(columns=columnas_salida)

    if aggfunc == "count":
        tabla = (
            df.groupby(["Nombre", "Rango_Hora"], dropna=False)[valor_col]
            .count()
            .reset_index(name="Valor")
        )
    else:
        df[valor_col] = pd.to_numeric(
            df[valor_col],
            errors="coerce"
        ).fillna(0)

        tabla = (
            df.groupby(["Nombre", "Rango_Hora"], dropna=False)[valor_col]
            .sum()
            .reset_index(name="Valor")
        )

    pivot = tabla.pivot_table(
        index="Nombre",
        columns="Rango_Hora",
        values="Valor",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    pivot = pivot.rename(columns={"Nombre": "NOMBRES"})

    for label in labels:
        if label not in pivot.columns:
            pivot[label] = 0

    pivot = pivot[["NOMBRES"] + labels].copy()

    pivot[total_col] = pivot[labels].sum(axis=1)

    meta_ppt = float(meta_ppt or 0)
    pivot["Meta PPT"] = meta_ppt

    pivot[pct_col] = 0.0

    if meta_ppt > 0:
        pivot[pct_col] = pivot[total_col] / meta_ppt

    pivot[fecha_col] = fecha_str

    pivot = pivot.sort_values(
        total_col,
        ascending=False
    ).reset_index(drop=True)

    if not pivot.empty:
        fila_total = {"NOMBRES": "TOTAL"}

        for label in labels:
            fila_total[label] = pivot[label].sum()

        fila_total[total_col] = pivot[total_col].sum()
        fila_total["Meta PPT"] = ""
        fila_total[pct_col] = ""
        fila_total[fecha_col] = fecha_str

        pivot = pd.concat(
            [pivot, pd.DataFrame([fila_total])],
            ignore_index=True
        )

    return pivot[columnas_salida]


def _generar_analisis_horas_cargue_despachos(eventos_despacho, fecha_str, ini, fin, meta_ppt_despachos):

    rangos = _generar_rangos_horarios_turno(fecha_str, ini, fin)
    labels = [r["label"] for r in rangos]

    columnas_salida = ["NOMBRES"] + labels + [
        "Total",
        "Meta PPT",
        "% Productividad",
        "Fecha"
    ]

    if eventos_despacho is None or eventos_despacho.empty:
        return pd.DataFrame(columns=columnas_salida)

    df = eventos_despacho.copy()

    df = _ensure_columns(df, [
        "Nombre",
        "Tipo_Evento_Despacho",
        "FechaHora"
    ])

    df = df[
        df["Tipo_Evento_Despacho"]
        .astype(str)
        .str.strip()
        .str.upper()
        .eq("CARGUE_PICK_REAL")
    ].copy()

    if df.empty:
        return pd.DataFrame(columns=columnas_salida)

    df["FechaHora_Analisis"] = pd.to_datetime(
        df["FechaHora"],
        errors="coerce"
    )

    df = df[df["FechaHora_Analisis"].notna()].copy()

    if df.empty:
        return pd.DataFrame(columns=columnas_salida)

    # Cada pick cuenta como 1 movimiento.
    df["Valor_Analisis"] = 1

    return _generar_matriz_horaria(
        df_base=df,
        fecha_str=fecha_str,
        ini=ini,
        fin=fin,
        meta_ppt=meta_ppt_despachos,
        valor_col="Valor_Analisis",
        aggfunc="sum",
        total_col="Total"
    )

def _generar_analisis_horas_tarea_bodega(
    movimientos,
    fecha_str,
    ini,
    fin,
    meta_ppt_bodega
):
    """
    Hoja: Analisis horas tarea Bodega

    Métrica:
    Cantidad de movimientos reales por hora por operario.

    Fuente:
    Movimientos consolidados con Area_Movimiento = Bodega.
    Se usa FH_inicio para ubicar el movimiento en la hora.
    """

    rangos = _generar_rangos_horarios_turno(fecha_str, ini, fin)
    labels = [r["label"] for r in rangos]

    columnas_salida = ["NOMBRES"] + labels + [
        "Total Movimientos",
        "Meta PPT",
        "% Productividad",
        "Fecha"
    ]

    if movimientos is None or movimientos.empty:
        return pd.DataFrame(columns=columnas_salida)

    df = movimientos.copy()

    df = _ensure_columns(df, [
        "Nombre",
        "Area_Movimiento",
        "Movimiento_ID",
        "FH_inicio"
    ])

    df = df[
        df["Area_Movimiento"]
        .astype(str)
        .str.strip()
        .str.upper()
        .eq("BODEGA")
    ].copy()

    if df.empty:
        return pd.DataFrame(columns=columnas_salida)

    df = df.drop_duplicates(
        subset=["Nombre", "Movimiento_ID"]
    ).copy()

    df["FechaHora_Analisis"] = pd.to_datetime(
        df["FH_inicio"],
        errors="coerce"
    )

    df["Valor_Analisis"] = df["Movimiento_ID"]

    return _generar_matriz_horaria(
        df_base=df,
        fecha_str=fecha_str,
        ini=ini,
        fin=fin,
        meta_ppt=meta_ppt_bodega,
        valor_col="Valor_Analisis",
        aggfunc="count",
        total_col="Total Movimientos"
    )


def _calcular_inactividad_bodega(movimientos, fecha_str, ini, fin, min_inactividad_reportable=25):
    """
    Calcula lapsos de inactividad para Bodega usando movimientos reales consolidados.

    Salida:
    - Incluye ID Operador.
    - No incluye Nombre_norm.
    """

    columnas = [
        "ID Operador",
        "Nombre",
        "Tipo_Inactividad",
        "Movimiento_Anterior_ID",
        "Movimiento_Siguiente_ID",
        "Fin_Tarea_Anterior",
        "Inicio_Tarea_Siguiente",
        "Tiempo_Inactivo",
        "Fecha"
    ]

    if movimientos is None or movimientos.empty:
        return pd.DataFrame(columns=columnas)

    df = movimientos.copy()

    df = _ensure_columns(df, [
        "ID Operador",
        "Nombre_norm",
        "Nombre",
        "Area_Movimiento",
        "Movimiento_ID",
        "FH_inicio",
        "FH_fin"
    ])

    df = df[
        df["Area_Movimiento"]
        .astype(str)
        .str.strip()
        .str.upper()
        .eq("BODEGA")
    ].copy()

    if df.empty:
        return pd.DataFrame(columns=columnas)

    df["FH_inicio"] = pd.to_datetime(df["FH_inicio"], errors="coerce")
    df["FH_fin"] = pd.to_datetime(df["FH_fin"], errors="coerce")

    # Si FH_fin viene vacío, se usa FH_inicio para no perder el movimiento.
    df["FH_fin"] = df["FH_fin"].fillna(df["FH_inicio"])

    df = df[
        df["FH_inicio"].notna() &
        df["FH_fin"].notna()
    ].copy()

    if df.empty:
        return pd.DataFrame(columns=columnas)

    fecha_base = datetime.strptime(fecha_str.strip(), "%d/%m/%Y").date()
    h_ini = datetime.strptime(ini, "%H:%M").time()
    h_fin = datetime.strptime(fin, "%H:%M").time()

    if h_ini <= h_fin:
        dt_ini_turno = datetime.combine(fecha_base, h_ini)
        dt_fin_turno = datetime.combine(fecha_base, h_fin)
    else:
        dt_ini_turno = datetime.combine(fecha_base - timedelta(days=1), h_ini)
        dt_fin_turno = datetime.combine(fecha_base, h_fin)

    registros = []

    group_cols = ["ID Operador", "Nombre_norm", "Nombre"]

    for _, g in df.groupby(group_cols, dropna=False):
        g = g.sort_values("FH_inicio").copy()

        id_operador = str(g["ID Operador"].iloc[0]).strip()
        nombre = str(g["Nombre"].iloc[0]).strip()

        primera = g.iloc[0]

        # ========================================================
        # 1. Inicio del turno a primera tarea
        # ========================================================
        gap_ini = (primera["FH_inicio"] - dt_ini_turno).total_seconds() / 60

        if gap_ini >= min_inactividad_reportable:
            registros.append({
                "ID Operador": id_operador,
                "Nombre": nombre,
                "Tipo_Inactividad": "INICIO_TURNO_A_PRIMERA_TAREA",
                "Movimiento_Anterior_ID": "",
                "Movimiento_Siguiente_ID": primera["Movimiento_ID"],
                "Fin_Tarea_Anterior": dt_ini_turno,
                "Inicio_Tarea_Siguiente": primera["FH_inicio"],
                "Tiempo_Inactivo": round(gap_ini, 2),
                "Fecha": fecha_str
            })

        # ========================================================
        # 2. Entre tareas de Bodega
        # ========================================================
        prev_fin = primera["FH_fin"]
        prev_mov_id = primera["Movimiento_ID"]

        for i in range(1, len(g)):
            row = g.iloc[i]

            inicio_actual = row["FH_inicio"]
            fin_actual = row["FH_fin"]
            mov_actual = row["Movimiento_ID"]

            gap = (inicio_actual - prev_fin).total_seconds() / 60

            if gap >= min_inactividad_reportable:
                registros.append({
                    "ID Operador": id_operador,
                    "Nombre": nombre,
                    "Tipo_Inactividad": "ENTRE_TAREAS_BODEGA",
                    "Movimiento_Anterior_ID": prev_mov_id,
                    "Movimiento_Siguiente_ID": mov_actual,
                    "Fin_Tarea_Anterior": prev_fin,
                    "Inicio_Tarea_Siguiente": inicio_actual,
                    "Tiempo_Inactivo": round(gap, 2),
                    "Fecha": fecha_str
                })

            # Si hay tareas solapadas o pegadas, conservar el fin más lejano.
            if fin_actual > prev_fin:
                prev_fin = fin_actual
                prev_mov_id = mov_actual

        # ========================================================
        # 3. Última tarea a fin del turno
        # ========================================================
        gap_fin = (dt_fin_turno - prev_fin).total_seconds() / 60

        if gap_fin >= min_inactividad_reportable:
            registros.append({
                "ID Operador": id_operador,
                "Nombre": nombre,
                "Tipo_Inactividad": "ULTIMA_TAREA_A_FIN_TURNO",
                "Movimiento_Anterior_ID": prev_mov_id,
                "Movimiento_Siguiente_ID": "",
                "Fin_Tarea_Anterior": prev_fin,
                "Inicio_Tarea_Siguiente": dt_fin_turno,
                "Tiempo_Inactivo": round(gap_fin, 2),
                "Fecha": fecha_str
            })

    if not registros:
        return pd.DataFrame(columns=columnas)

    resultado = pd.DataFrame(registros)

    resultado = resultado.sort_values([
        "Nombre",
        "Fin_Tarea_Anterior",
        "Inicio_Tarea_Siguiente"
    ]).reset_index(drop=True)

    return resultado[columnas]


# ============================================================
# FORMATO EXCEL PARA HOJAS DE ANÁLISIS
# ============================================================

def _formatear_hoja_analisis_horas(writer, sheet_name):
    """
    Aplica formato básico a las hojas tipo:
    NOMBRES | HORAS | Total | Meta PPT | % PRODUCTIVIDAD | FECHA
    """
    try:
        wb = writer.book
        ws = wb[sheet_name]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        fill_header = PatternFill("solid", fgColor="156082")
        font_header = Font(color="FFFFFF", bold=True)
        fill_total = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(border_style="thin", color="808080")

        max_row = ws.max_row
        max_col = ws.max_column

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        for row in range(2, max_row + 1):
            if str(ws.cell(row=row, column=1).value).strip().upper() == "TOTAL":
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = fill_total
                    ws.cell(row=row, column=col).font = Font(bold=True)

        for col in range(1, max_col + 1):
            header = str(ws.cell(row=1, column=col).value).strip().upper()

            if header == "% PRODUCTIVIDAD":
                for row in range(2, max_row + 1):
                    ws.cell(row=row, column=col).number_format = "0.00%"

        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            header = str(ws.cell(row=1, column=col).value or "")

            if col == 1:
                ws.column_dimensions[letter].width = 32
            elif header.upper() in [
                "TOTAL",
                "TOTAL MOVIMIENTOS",
                "META PPT",
                "% PRODUCTIVIDAD",
                "FECHA"
            ]:
                ws.column_dimensions[letter].width = 18
            else:
                ws.column_dimensions[letter].width = 14

        ws.freeze_panes = "B2"

    except Exception:
        pass
    
def _aplicar_estilo_excel_resultados(writer):
    """
    Aplica estilo general al Excel resultado sin crear tablas internas.

    Incluye:
    - filtros en encabezados
    - encabezado con color
    - congelar fila superior
    - bordes
    - ajuste de anchos
    - formato de porcentajes
    - barras de productividad en % Productividad
    - color de fondo en horas muertas con valor 0
    - formato HH:MM para Tiempo_Inactivo
    - ocultar cuadrículas

    Nota:
    No se crean objetos Table de Excel para evitar errores de reparación
    como /xl/tables/table1.xml.
    """

    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import DataBarRule

        wb = writer.book

        fill_header = PatternFill("solid", fgColor="156082")
        font_header = Font(color="FFFFFF", bold=True)

        fill_total = PatternFill("solid", fgColor="D9EAF7")

        # Color para horas muertas / celdas en 0.
        fill_cero = PatternFill("solid", fgColor="FFC7CE")
        font_cero = Font(color="9C0006")

        thin = Side(border_style="thin", color="B7B7B7")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False

            max_row = ws.max_row
            max_col = ws.max_column

            if max_row < 1 or max_col < 1:
                continue

            last_col_letter = get_column_letter(max_col)

            # Congelar encabezado
            ws.freeze_panes = "A2"

            # Filtro automático seguro
            ws.auto_filter.ref = f"A1:{last_col_letter}{max_row}"

            # ====================================================
            # Estilo encabezado
            # ====================================================
            for cell in ws[1]:
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                cell.border = border

            # ====================================================
            # Estilo cuerpo
            # ====================================================
            for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=False
                    )

            # Primera columna alineada a la izquierda
            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=False
                    )

            # ====================================================
            # Identificar columnas por encabezado
            # ====================================================
            col_por_header = {}

            for col_idx in range(1, max_col + 1):
                header = str(ws.cell(row=1, column=col_idx).value or "").strip()
                col_por_header[header.upper()] = col_idx

            # ====================================================
            # Resaltar filas TOTAL
            # ====================================================
            for row_idx in range(2, max_row + 1):
                valor_col_a = str(ws.cell(row=row_idx, column=1).value or "").strip().upper()
                valor_col_b = str(ws.cell(row=row_idx, column=2).value or "").strip().upper()

                if valor_col_a == "TOTAL" or valor_col_b == "TOTAL":
                    for col_idx in range(1, max_col + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = fill_total
                        cell.font = Font(bold=True)
                        cell.border = border

            # ====================================================
            # Colorear horas muertas en hojas de análisis horario
            # ====================================================
            titulo_hoja = ws.title.upper()
            es_hoja_analisis_horas = (
                "ANALISIS" in titulo_hoja
                and (
                    "CARGUES" in titulo_hoja
                    or "MOVIMIENTOS" in titulo_hoja
                    or "BODEGA" in titulo_hoja
                    or "DESPACHOS" in titulo_hoja
                )
            )

            if es_hoja_analisis_horas:
                for col_idx in range(1, max_col + 1):
                    header = str(ws.cell(row=1, column=col_idx).value or "").strip()
                    header_upper = header.upper()

                    # Las columnas horarias tienen formato tipo "07:00 a 08:00"
                    es_columna_hora = " A " in header_upper

                    if not es_columna_hora:
                        continue

                    for row_idx in range(2, max_row + 1):
                        nombre_fila = str(ws.cell(row=row_idx, column=1).value or "").strip().upper()

                        # No colorear la fila TOTAL
                        if nombre_fila == "TOTAL":
                            continue

                        cell = ws.cell(row=row_idx, column=col_idx)

                        try:
                            valor = float(cell.value or 0)
                        except Exception:
                            valor = None

                        if valor == 0:
                            cell.fill = fill_cero
                            cell.font = font_cero

            # ====================================================
            # Formato por encabezado
            # ====================================================
            for col_idx in range(1, max_col + 1):
                header = str(ws.cell(row=1, column=col_idx).value or "").strip()
                header_upper = header.upper()

                for row_idx in range(2, max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)

                    if "%" in header_upper or "PORCENTAJE" in header_upper:
                        cell.number_format = "0%"

                    elif header_upper == "FECHA":
                        cell.number_format = "dd/mm/yyyy"
                    
                    elif (
                        "FH_" in header_upper
                        or "INICIO" in header_upper
                        or "FIN" in header_upper
                    ):
                        cell.number_format = "dd/mm/yyyy hh:mm"

                    elif "DURACION" in header_upper:
                        cell.number_format = "0.00"

            # ====================================================
            # Barras rojas para % Productividad
            # ====================================================
            col_productividad = None

            for col_idx in range(1, max_col + 1):
                header = str(ws.cell(row=1, column=col_idx).value or "").strip().upper()

                if header in ["% PRODUCTIVIDAD", "% PRODUCTIVIDAD "]:
                    col_productividad = col_idx
                    break

            if col_productividad is not None and max_row >= 2:
                letra = get_column_letter(col_productividad)
                rango = f"{letra}2:{letra}{max_row}"

                regla_barra = DataBarRule(
                    start_type="num",
                    start_value=0,
                    end_type="max",
                    color="FF6666",
                    showValue=True
                )

                ws.conditional_formatting.add(rango, regla_barra)

                for row_idx in range(2, max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_productividad)
                    cell.number_format = "0%"

                    # Borde rojo similar al ejemplo
                    cell.border = Border(
                        top=Side(border_style="thin", color="FF6666"),
                        left=Side(border_style="thin", color="FF6666"),
                        right=Side(border_style="thin", color="FF6666"),
                        bottom=Side(border_style="thin", color="FF6666")
                    )

            # ====================================================
            # Formato HH:MM acumulado para tiempos en minutos
            # ====================================================
            columnas_tiempo_minutos = [
                "TIEMPO INACTIVO",
                "TIEMPO_INACTIVO",
                "TIEMPO DURACION",
                "TIEMPO_DURACION",
                "TIEMPO DURACIÓN",
            ]

            for col_idx in range(1, max_col + 1):
                header = str(ws.cell(row=1, column=col_idx).value or "").strip().upper()

                if header not in columnas_tiempo_minutos:
                    continue

                for row_idx in range(2, max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)

                    try:
                        if cell.value is None or str(cell.value).strip() == "":
                            continue

                        minutos = float(cell.value)

                        # Excel maneja horas como fracción de día.
                        # 60 minutos = 60 / 1440 = 01:00
                        cell.value = minutos / 1440

                        # Formato acumulado: permite más de 24 horas.
                        # Ejemplo: 1500 minutos = 25:00
                        cell.number_format = "[h]:mm"

                    except Exception:
                        pass

            # ====================================================
            # Ajuste de ancho de columnas
            # ====================================================
            for col_idx in range(1, max_col + 1):
                letter = get_column_letter(col_idx)
                header = str(ws.cell(row=1, column=col_idx).value or "").strip()
                header_upper = header.upper()

                max_len = len(header)

                for row_idx in range(2, min(max_row, 250) + 1):
                    valor = ws.cell(row=row_idx, column=col_idx).value

                    if valor is None:
                        continue

                    max_len = max(max_len, len(str(valor)))

                ancho = min(max(max_len + 2, 12), 45)

                if header_upper in [
                    "NOMBRE",
                    "NOMBRES",
                    "RESUMEN_MOVIMIENTO",
                    "OBSERVACION_LOGICA"
                ]:
                    ancho = min(max(ancho, 28), 55)

                if header_upper in [
                    "TOTAL",
                    "TOTAL MOVIMIENTOS",
                    "META PPT",
                    "% PRODUCTIVIDAD",
                    "FECHA"
                ]:
                    ancho = max(ancho, 18)

                if header_upper in [
                    "MINUTOS_INACTIVOS",
                    "MINUTOS INACTIVOS",
                    "TIEMPO INACTIVO",
                    "TIEMPO_INACTIVO",
                    "TIEMPO DURACION",
                    "TIEMPO_DURACION",
                    "TIEMPO DURACIÓN",
                    "TIEMPO_DURACIÓN"
                ]:
                    ancho = max(ancho, 18)

                ws.column_dimensions[letter].width = ancho

    except Exception:
        # El formato no debe detener el procesamiento.
        pass
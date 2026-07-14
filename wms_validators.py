import os
from datetime import datetime, timedelta, date
from openpyxl import load_workbook

from wms_utils import _parse_hora_excel


MAX_DIAS_RANGO_EXCEL = 31


def _parse_fecha_wms_mmdd(valor):
    """
    Convierte valores de fecha del Excel WMS a date.

    El WMS descarga la fecha en formato:
        MM/DD/YYYY

    Ejemplo:
        06/09/2026 = 9 de junio de 2026
        06/10/2026 = 10 de junio de 2026

    Esto evita que el software interprete erróneamente:
        06/09/2026 como 6 de septiembre de 2026.
    """
    if valor is None:
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()

    if not texto or texto.lower() == "nan":
        return None

    texto_corto = texto[:10]

    formatos = [
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y"
    ]

    for fmt in formatos:
        try:
            return datetime.strptime(texto_corto, fmt).date()
        except Exception:
            pass

    return None


def validar_fecha_y_turno_excel_rapido(
    ruta_excel,
    fecha_seleccionada,
    hora_inicio,
    hora_fin,
    columna_fecha="Fecha",
    columna_tiempo="Tiempo"
):
    """
    Valida rápidamente:
    1. Que la fecha seleccionada esté dentro del rango del Excel.
    2. Que existan registros dentro del turno seleccionado.

    Usa openpyxl read_only para no cargar todo el Excel con pandas.

    Importante:
    - La fecha seleccionada desde la interfaz se mantiene en formato dd/mm/aaaa.
    - La fecha del Excel WMS se interpreta como MM/DD/YYYY.
    """

    wb = None

    try:
        try:
            fecha_base = datetime.strptime(
                str(fecha_seleccionada).strip(),
                "%d/%m/%Y"
            ).date()
        except Exception:
            return (
                False,
                "La fecha de operación no es válida.\n\nUsa el formato dd/mm/aaaa.",
                None,
                None,
                0
            )

        try:
            h_ini = datetime.strptime(str(hora_inicio).strip(), "%H:%M").time()
            h_fin = datetime.strptime(str(hora_fin).strip(), "%H:%M").time()
        except Exception:
            return (
                False,
                "La hora de inicio o fin no es válida.\n\nUsa el formato HH:MM, por ejemplo 07:30 o 19:30.",
                None,
                None,
                0
            )

        wb = load_workbook(
            filename=ruta_excel,
            read_only=True,
            data_only=True
        )

        ws = wb.active

        encabezados = []

        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            encabezados.append(
                str(cell.value).strip() if cell.value is not None else ""
            )

        idx_fecha = None
        idx_tiempo = None

        for idx, nombre_col in enumerate(encabezados, start=1):
            nombre_normalizado = nombre_col.strip().lower()

            if nombre_normalizado == columna_fecha.strip().lower():
                idx_fecha = idx

            if nombre_normalizado == columna_tiempo.strip().lower():
                idx_tiempo = idx

        if idx_fecha is None:
            return (
                False,
                f"No se encontró la columna '{columna_fecha}' en el Excel origen.",
                None,
                None,
                0
            )

        if idx_tiempo is None:
            return (
                False,
                f"No se encontró la columna '{columna_tiempo}' en el Excel origen.",
                None,
                None,
                0
            )

        fecha_min = None
        fecha_max = None
        fechas_validas = 0
        registros_turno = 0

        cruza_medianoche = h_ini > h_fin

        min_col = min(idx_fecha, idx_tiempo)
        max_col = max(idx_fecha, idx_tiempo)

        pos_fecha = idx_fecha - min_col
        pos_tiempo = idx_tiempo - min_col

        for row in ws.iter_rows(
            min_row=2,
            min_col=min_col,
            max_col=max_col,
            values_only=True
        ):
            valor_fecha = row[pos_fecha]
            valor_tiempo = row[pos_tiempo]

            fecha_actual = _parse_fecha_wms_mmdd(valor_fecha)

            if fecha_actual is None:
                continue

            fechas_validas += 1

            if fecha_min is None or fecha_actual < fecha_min:
                fecha_min = fecha_actual

            if fecha_max is None or fecha_actual > fecha_max:
                fecha_max = fecha_actual

            hora_actual = _parse_hora_excel(valor_tiempo)

            if hora_actual is None:
                continue

            if not cruza_medianoche:
                if (
                    fecha_actual == fecha_base
                    and hora_actual >= h_ini
                    and hora_actual <= h_fin
                ):
                    registros_turno += 1
            else:
                fecha_inicio_turno = fecha_base - timedelta(days=1)

                cond_noche = (
                    fecha_actual == fecha_inicio_turno
                    and hora_actual >= h_ini
                )

                cond_madrugada = (
                    fecha_actual == fecha_base
                    and hora_actual <= h_fin
                )

                if cond_noche or cond_madrugada:
                    registros_turno += 1

        if fechas_validas == 0 or fecha_min is None or fecha_max is None:
            return (
                False,
                f"La columna '{columna_fecha}' no contiene fechas válidas.",
                None,
                None,
                0
            )

        if fecha_base < fecha_min or fecha_base > fecha_max:
            mensaje = (
                "La fecha de operación seleccionada está fuera del rango del Excel origen.\n\n"
                f"Fecha seleccionada: {fecha_base.strftime('%d/%m/%Y')}\n"
                f"Rango disponible en el archivo: {fecha_min.strftime('%d/%m/%Y')} hasta {fecha_max.strftime('%d/%m/%Y')}\n\n"
                "Seleccione una fecha dentro del rango disponible para continuar."
            )

            return False, mensaje, fecha_min, fecha_max, registros_turno

        if registros_turno == 0:
            if not cruza_medianoche:
                mensaje_turno = (
                    "La fecha existe en el Excel, pero no hay registros para el turno seleccionado.\n\n"
                    f"Fecha operación: {fecha_base.strftime('%d/%m/%Y')}\n"
                    f"Turno seleccionado: {h_ini.strftime('%H:%M')} a {h_fin.strftime('%H:%M')}\n\n"
                    "Verifica la hora de inicio, la hora fin o la fecha de operación."
                )
            else:
                fecha_inicio_turno = fecha_base - timedelta(days=1)

                mensaje_turno = (
                    "La fecha existe en el Excel, pero no hay registros para el turno nocturno seleccionado.\n\n"
                    f"Fecha operación seleccionada: {fecha_base.strftime('%d/%m/%Y')}\n"
                    f"Turno nocturno evaluado: {fecha_inicio_turno.strftime('%d/%m/%Y')} {h_ini.strftime('%H:%M')} "
                    f"hasta {fecha_base.strftime('%d/%m/%Y')} {h_fin.strftime('%H:%M')}\n\n"
                    "Verifica si la fecha de operación corresponde al día de cierre del turno nocturno."
                )

            return False, mensaje_turno, fecha_min, fecha_max, registros_turno

        mensaje_ok = (
            "Fecha y turno validados correctamente.\n\n"
            f"Fecha operación: {fecha_base.strftime('%d/%m/%Y')}\n"
            f"Turno: {h_ini.strftime('%H:%M')} a {h_fin.strftime('%H:%M')}\n"
            f"Registros detectados en el turno: {registros_turno}\n"
            f"Rango disponible en archivo: {fecha_min.strftime('%d/%m/%Y')} hasta {fecha_max.strftime('%d/%m/%Y')}"
        )

        return True, mensaje_ok, fecha_min, fecha_max, registros_turno

    except Exception as e:
        return (
            False,
            f"Error validando fecha y turno del Excel origen:\n\n{str(e)}",
            None,
            None,
            0
        )

    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


def analizar_excel_origen_rapido(
    ruta_excel,
    columna_fecha="Fecha",
    max_dias_rango=MAX_DIAS_RANGO_EXCEL
):
    """
    Analiza rápidamente el Excel origen leyendo únicamente la columna Fecha.

    Importante:
    - La columna Fecha del WMS se interpreta como MM/DD/YYYY.
    """

    wb = None

    try:
        if not ruta_excel or not str(ruta_excel).strip():
            return {
                "ok": False,
                "mensaje": "No se encontró la ruta del Excel origen.",
                "fecha_min": None,
                "fecha_max": None,
                "total_fechas_validas": 0,
                "dias_rango": 0,
                "fecha_sugerida": None,
                "bloqueado_por_rango": False
            }

        ruta_excel = str(ruta_excel).strip()

        if not os.path.exists(ruta_excel):
            return {
                "ok": False,
                "mensaje": "El archivo Excel origen no existe.",
                "fecha_min": None,
                "fecha_max": None,
                "total_fechas_validas": 0,
                "dias_rango": 0,
                "fecha_sugerida": None,
                "bloqueado_por_rango": False
            }

        wb = load_workbook(
            filename=ruta_excel,
            read_only=True,
            data_only=True
        )

        ws = wb.active

        encabezados = []

        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            encabezados.append(
                str(cell.value).strip() if cell.value is not None else ""
            )

        idx_fecha = None

        for idx, nombre_col in enumerate(encabezados, start=1):
            if nombre_col.strip().lower() == columna_fecha.strip().lower():
                idx_fecha = idx
                break

        if idx_fecha is None:
            return {
                "ok": False,
                "mensaje": f"No se encontró la columna '{columna_fecha}' en el Excel origen.",
                "fecha_min": None,
                "fecha_max": None,
                "total_fechas_validas": 0,
                "dias_rango": 0,
                "fecha_sugerida": None,
                "bloqueado_por_rango": False
            }

        fecha_min = None
        fecha_max = None
        total_validas = 0

        for row in ws.iter_rows(
            min_row=2,
            min_col=idx_fecha,
            max_col=idx_fecha,
            values_only=True
        ):
            fecha_actual = _parse_fecha_wms_mmdd(row[0])

            if fecha_actual is None:
                continue

            total_validas += 1

            if fecha_min is None or fecha_actual < fecha_min:
                fecha_min = fecha_actual

            if fecha_max is None or fecha_actual > fecha_max:
                fecha_max = fecha_actual

        if total_validas == 0 or fecha_min is None or fecha_max is None:
            return {
                "ok": False,
                "mensaje": f"La columna '{columna_fecha}' no contiene fechas válidas.",
                "fecha_min": None,
                "fecha_max": None,
                "total_fechas_validas": 0,
                "dias_rango": 0,
                "fecha_sugerida": None,
                "bloqueado_por_rango": False
            }

        dias_rango = (fecha_max - fecha_min).days + 1

        if dias_rango > max_dias_rango:
            return {
                "ok": False,
                "mensaje": (
                    "El Excel origen contiene un rango de fechas demasiado amplio.\n\n"
                    f"Rango detectado: {fecha_min.strftime('%d/%m/%Y')} a {fecha_max.strftime('%d/%m/%Y')}\n"
                    f"Días detectados: {dias_rango}\n"
                    f"Máximo permitido: {max_dias_rango} días\n\n"
                    "Para evitar saturación, carga un Excel origen filtrado con un rango menor de fechas."
                ),
                "fecha_min": fecha_min,
                "fecha_max": fecha_max,
                "total_fechas_validas": total_validas,
                "dias_rango": dias_rango,
                "fecha_sugerida": None,
                "bloqueado_por_rango": True
            }

        fecha_sugerida = fecha_max

        return {
            "ok": True,
            "mensaje": (
                "Excel origen validado correctamente.\n\n"
                f"Rango detectado: {fecha_min.strftime('%d/%m/%Y')} a {fecha_max.strftime('%d/%m/%Y')}\n"
                f"Días detectados: {dias_rango}\n"
                f"Fecha operación sugerida: {fecha_sugerida.strftime('%d/%m/%Y')}"
            ),
            "fecha_min": fecha_min,
            "fecha_max": fecha_max,
            "total_fechas_validas": total_validas,
            "dias_rango": dias_rango,
            "fecha_sugerida": fecha_sugerida,
            "bloqueado_por_rango": False
        }

    except Exception as e:
        return {
            "ok": False,
            "mensaje": f"Error analizando el Excel origen:\n\n{str(e)}",
            "fecha_min": None,
            "fecha_max": None,
            "total_fechas_validas": 0,
            "dias_rango": 0,
            "fecha_sugerida": None,
            "bloqueado_por_rango": False
        }

    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


def validar_excel_y_autoseleccionar_fecha_ui(
    window,
    ruta_excel,
    cache_fechas,
    max_dias_rango=MAX_DIAS_RANGO_EXCEL
):
    """
    Valida el Excel apenas se carga.
    Si el rango es permitido, selecciona automáticamente una fecha operación válida.
    Si el rango es demasiado amplio, bloquea visualmente el proceso.
    """

    resultado = analizar_excel_origen_rapido(
        ruta_excel=ruta_excel,
        columna_fecha="Fecha",
        max_dias_rango=max_dias_rango
    )

    cache_fechas["ruta"] = ruta_excel
    cache_fechas["fecha_min"] = resultado.get("fecha_min")
    cache_fechas["fecha_max"] = resultado.get("fecha_max")
    cache_fechas["total_validas"] = resultado.get("total_fechas_validas", 0)
    cache_fechas["dias_rango"] = resultado.get("dias_rango", 0)
    cache_fechas["bloqueado_por_rango"] = resultado.get("bloqueado_por_rango", False)
    cache_fechas["excel_ok"] = resultado.get("ok", False)

    if not resultado["ok"]:
        if resultado.get("bloqueado_por_rango"):
            window["-FECHA_ESTADO-"].update(
                f"❌ Excel bloqueado: rango de {resultado['dias_rango']} días. Máximo permitido: {max_dias_rango}.",
                text_color="#b00020"
            )
        else:
            window["-FECHA_ESTADO-"].update(
                f"❌ {resultado['mensaje'][:140]}",
                text_color="#b00020"
            )

        return False, resultado

    fecha_sugerida = resultado["fecha_sugerida"]

    if fecha_sugerida is not None:
        window["-FECHA-"].update(fecha_sugerida.strftime("%d/%m/%Y"))

    window["-FECHA_ESTADO-"].update(
        f"✅ Excel válido. Fecha operación seleccionada: {fecha_sugerida.strftime('%d/%m/%Y')} | "
        f"Rango: {resultado['fecha_min'].strftime('%d/%m/%Y')} a {resultado['fecha_max'].strftime('%d/%m/%Y')}",
        text_color="#007a33"
    )

    return True, resultado


def validar_fecha_ui_inmediata(window, ruta_excel, fecha_str, cache_fechas):
    """
    Valida inmediatamente la fecha seleccionada contra el rango real del Excel origen.
    Usa cache para no releer el Excel cada vez.

    La fecha que escribe el usuario en la UI sigue siendo dd/mm/aaaa.
    """

    try:
        if not ruta_excel or not str(ruta_excel).strip():
            window["-FECHA_ESTADO-"].update(
                "⚠ Selecciona primero el Excel origen.",
                text_color="#b36b00"
            )
            return False

        ruta_excel = str(ruta_excel).strip()

        if not os.path.exists(ruta_excel):
            window["-FECHA_ESTADO-"].update(
                "❌ El archivo Excel origen no existe.",
                text_color="#b00020"
            )
            return False

        fecha_min = cache_fechas.get("fecha_min")
        fecha_max = cache_fechas.get("fecha_max")

        if fecha_min is None or fecha_max is None:
            window["-FECHA_ESTADO-"].update(
                "⚠ Carga primero un Excel origen válido.",
                text_color="#b36b00"
            )
            return False

        try:
            fecha_usuario = datetime.strptime(
                str(fecha_str).strip(),
                "%d/%m/%Y"
            ).date()
        except Exception:
            window["-FECHA_ESTADO-"].update(
                "❌ Fecha inválida. Usa formato dd/mm/aaaa.",
                text_color="#b00020"
            )
            return False

        if fecha_usuario < fecha_min or fecha_usuario > fecha_max:
            window["-FECHA_ESTADO-"].update(
                f"❌ Fecha fuera de rango. Disponible: {fecha_min.strftime('%d/%m/%Y')} a {fecha_max.strftime('%d/%m/%Y')}",
                text_color="#b00020"
            )
            return False

        if cache_fechas.get("bloqueado_por_rango"):
            window["-FECHA_ESTADO-"].update(
                f"❌ Excel bloqueado: rango de {cache_fechas.get('dias_rango', 0)} días. Máximo permitido: {MAX_DIAS_RANGO_EXCEL}.",
                text_color="#b00020"
            )
            return False

        window["-FECHA_ESTADO-"].update(
            f"✅ Fecha válida. Rango del Excel: {fecha_min.strftime('%d/%m/%Y')} a {fecha_max.strftime('%d/%m/%Y')}",
            text_color="#007a33"
        )

        return True

    except Exception as e:
        window["-FECHA_ESTADO-"].update(
            f"❌ Error validando fecha: {str(e)[:120]}",
            text_color="#b00020"
        )
        return False